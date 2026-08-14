"""Deterministic Excel inspection, query and export tools.

No LLM or orchestration logic belongs here. Tool responses are deliberately compact so
that an external agent can decide the next call without receiving the workbook itself.
"""
from __future__ import annotations

import csv
import json
import math
import os
import re
import uuid
from collections import Counter, deque
from collections.abc import Iterable, Iterator
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .sessions import ARTIFACT_RE, session_file
from .tools import ToolError, tool

MAX_PREVIEW_ROWS = int(os.getenv("MAX_PREVIEW_ROWS", "100"))
MAX_QUERY_PREVIEW_ROWS = int(os.getenv("MAX_QUERY_PREVIEW_ROWS", "100"))
MAX_DISTINCT_VALUES = 500
SUPPORTED_OPENPYXL_SUFFIXES = {".xlsx", ".xlsm", ".xltx", ".xltm"}
SUPPORTED_SUFFIXES = SUPPORTED_OPENPYXL_SUFFIXES | {".xls"}


def _schema(name: str, description: str, properties: dict[str, Any], required: list[str] | None = None) -> dict[str, Any]:
    return {
        "type": "function",
        "function": {
            "name": name,
            "description": description,
            "parameters": {
                "type": "object",
                "properties": properties,
                "required": required or [],
                "additionalProperties": False,
            },
        },
    }


def _id(prefix: str) -> str:
    return f"{prefix}_{uuid.uuid4().hex[:16]}"


def _opaque_id(value: Any, prefix: str) -> str:
    """Validate and normalize an opaque ID returned by a previous tool.

    Tool-use models occasionally preserve an invisible trailing newline or space
    while copying IDs from a JSON response.  IDs are server-generated, never user
    supplied identifiers, so trimming surrounding whitespace is safe and avoids a
    needless failed/retried tool iteration.  The strict pattern still prevents an
    arbitrary key or path from being used for a state lookup.
    """
    if not isinstance(value, str):
        raise ToolError("INVALID_ARGUMENTS", f"{prefix}_id must be a string")
    normalized = value.strip()
    pattern = re.compile(rf"^{re.escape(prefix)}_[A-Za-z0-9_-]{{8,64}}$")
    if not pattern.fullmatch(normalized):
        raise ToolError("INVALID_ARGUMENTS", f"Invalid {prefix}_id")
    return normalized


def _file_path(state: dict[str, Any]) -> Path:
    stored_path = state.get("file_path")
    if not isinstance(stored_path, str):
        raise ToolError("SESSION_FILE_MISSING", "Session input file is unavailable")
    try:
        path = session_file(state["session_id"], stored_path)
    except ValueError as error:
        raise ToolError("SESSION_FILE_MISSING", "Session input file is unavailable") from error
    if not path.is_file():
        raise ToolError("SESSION_FILE_MISSING", "Session input file is unavailable")
    return path


def _suffix(state: dict[str, Any]) -> str:
    suffix = Path(str(state.get("file_path", ""))).suffix.lower()
    if suffix not in SUPPORTED_SUFFIXES:
        raise ToolError("UNSUPPORTED_FILE_TYPE", "Only .xlsx, .xlsm, .xltx and .xls Excel files are supported")
    return suffix


def _json_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped or stripped.casefold() in _EMPTY_PLACEHOLDERS:
            return None
        return value
    if isinstance(value, (int, bool)):
        return value
    if isinstance(value, float):
        return None if math.isnan(value) or math.isinf(value) else value
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    # numpy/pandas scalar and Decimal support, without hard importing those packages.
    if hasattr(value, "item"):
        try:
            return _json_value(value.item())
        except (ValueError, TypeError):
            pass
    return str(value)


_EMPTY_PLACEHOLDERS = frozenset({"…", "...", "⋯", "n/a", "null"})
_NOTE_SHEET_NAMES = frozenset(
    {"description", "notes", "note", "readme", "contents", "afosheet", "changelog"}
)
_PREAMBLE_SHEET_NAMES = frozenset({"index"})
_GENERIC_COLUMN_TOKENS = frozenset(
    {"rate", "index", "price", "value", "total", "average", "avg", "percent", "bbl", "mmbtu"}
)
_MISSING_QUERY_RE = re.compile(
    r"\b(?:missing|ellipsis|unavailable|empty|null|n/a|нет\s+данн|пуст)\b",
    re.IGNORECASE,
)
_LATEST_COUNT_RE = re.compile(r"(?:latest|last|последн\w*)\s+(\d+)", re.IGNORECASE)
_CAP_COUNT_RE = re.compile(r"(?:up to|at most|no more than|не более|до)\s+(\d+)", re.IGNORECASE)
_LATEST_PERIOD_RE = re.compile(
    r"(?:latest|last)(?: available)? (?:period|observation)(?: only)?",
    re.IGNORECASE,
)
_UNIT_HINT = re.compile(
    r"(?:[\$%/()]|index|\bbbl\b|\bmmbtu\b|\bmt\b|\bkg\b|\btonne|\b2010\s*=\s*100)",
    re.IGNORECASE,
)
_SHEET_SUFFIX_ALIASES = {
    "p": ("people", "persons", "person"),
    "m": ("men", "male", "males", "man"),
    "w": ("women", "female", "females", "woman"),
}
_MATCH_STOP_WORDS = frozenset(
    {
        "extract", "from", "this", "workbook", "data", "table", "sheet", "file", "excel", "xlsx",
        "csv", "all", "with", "and", "the", "for", "not", "main", "visible", "orders", "order",
        "return", "latest", "last", "only", "ignore", "use", "available", "period", "rows", "row",
        "values", "value", "give", "me", "what", "is", "are", "please", "find", "shown", "those",
        "using", "into", "onto", "that", "this", "these", "them", "then", "than", "also", "just",
        "извлеки", "извлечь", "из", "этого", "этой", "этот", "книги", "книга", "файла", "файл",
        "данные", "таблица", "лист", "все", "с", "и", "не", "главной", "видимого",
    }
)


def _is_empty(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        stripped = value.strip()
        return not stripped or stripped.casefold() in _EMPTY_PLACEHOLDERS
    return False


def _nonempty_count(row: Iterable[Any]) -> int:
    return sum(not _is_empty(value) for value in row)


def _trim_row(row: tuple[Any, ...] | list[Any], max_col: int) -> list[Any]:
    values = list(row[:max_col])
    while values and _is_empty(values[-1]):
        values.pop()
    return values


def _xlsx_sheet_names_and_dimensions(path: Path) -> list[dict[str, Any]]:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    except Exception as error:
        raise ToolError("WORKBOOK_READ_ERROR", "Cannot read Excel workbook") from error
    try:
        return [
            {
                "name": worksheet.title,
                "max_row": worksheet.max_row or 0,
                "max_column": worksheet.max_column or 0,
                "state": worksheet.sheet_state,
            }
            for worksheet in workbook.worksheets
        ]
    finally:
        workbook.close()


def _xls_frames(path: Path) -> dict[str, Any]:
    try:
        import pandas as pd

        with pd.ExcelFile(path, engine="calamine") as excel_file:
            return {
                name: pd.read_excel(excel_file, sheet_name=name, header=None, dtype=object)
                for name in excel_file.sheet_names
            }
    except Exception as error:
        raise ToolError("WORKBOOK_READ_ERROR", "Cannot read .xls workbook") from error


def _sheets(state: dict[str, Any]) -> list[dict[str, Any]]:
    path = _file_path(state)
    if _suffix(state) in SUPPORTED_OPENPYXL_SUFFIXES:
        return _xlsx_sheet_names_and_dimensions(path)
    frames = _xls_frames(path)
    return [
        {"name": name, "max_row": int(frame.shape[0]), "max_column": int(frame.shape[1]), "state": "visible"}
        for name, frame in frames.items()
    ]


def _ensure_sheet(state: dict[str, Any], sheet: str) -> None:
    available = [item["name"] for item in _sheets(state)]
    if sheet not in available:
        raise ToolError("SHEET_NOT_FOUND", f"Sheet {sheet} not found", {"available_sheets": available})


def _iter_sheet_rows(state: dict[str, Any], sheet: str, min_row: int = 1, max_row: int | None = None) -> Iterator[tuple[int, list[Any]]]:
    path = _file_path(state)
    suffix = _suffix(state)
    if suffix in SUPPORTED_OPENPYXL_SUFFIXES:
        try:
            workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
            if sheet not in workbook.sheetnames:
                raise ToolError("SHEET_NOT_FOUND", f"Sheet {sheet} not found", {"available_sheets": workbook.sheetnames})
            worksheet = workbook[sheet]
            final_row = max_row if max_row is not None else worksheet.max_row
            for index, row in enumerate(
                worksheet.iter_rows(min_row=min_row, max_row=final_row, values_only=True), start=min_row
            ):
                yield index, list(row)
        except ToolError:
            raise
        except Exception as error:
            raise ToolError("WORKBOOK_READ_ERROR", "Cannot read Excel workbook") from error
        finally:
            if "workbook" in locals():
                workbook.close()
        return

    frames = _xls_frames(path)
    if sheet not in frames:
        raise ToolError("SHEET_NOT_FOUND", f"Sheet {sheet} not found", {"available_sheets": list(frames)})
    frame = frames[sheet]
    end = min(max_row or len(frame), len(frame))
    for index in range(max(1, min_row), end + 1):
        yield index, [_json_value(value) for value in frame.iloc[index - 1].tolist()]


def _headers(header_values: list[Any], max_col: int) -> list[str]:
    result: list[str] = []
    used: Counter[str] = Counter()
    for index in range(max_col):
        raw = header_values[index] if index < len(header_values) else None
        base = str(raw).strip() if not _is_empty(raw) else f"Column {get_column_letter(index + 1)}"
        used[base] += 1
        result.append(base if used[base] == 1 else f"{base} ({used[base]})")
    return result


def _row_bounds(row: list[Any]) -> tuple[int, int] | None:
    """Return the 1-based first/last meaningful cells in a row."""
    nonempty = [index for index, value in enumerate(row, start=1) if not _is_empty(value)]
    if not nonempty:
        return None
    return min(nonempty), max(nonempty)


def _filled_header_values(row: list[Any], start_column: int, end_column: int) -> list[Any]:
    """Fill merged-header blanks horizontally for an OpenAI-friendly display name.

    `openpyxl` deliberately exposes non-anchor cells of a merged range as `None` in
    read-only mode. For ``Q1`` merged over two sub-columns this produces
    ``["Q1", None]``. Carrying the anchor to the right gives useful deterministic
    names while still retaining the physical table coordinates.
    """
    result: list[Any] = []
    last_value: Any = None
    for index in range(start_column - 1, end_column):
        value = row[index] if index < len(row) else None
        if not _is_empty(value):
            last_value = value
        result.append(value if not _is_empty(value) else last_value)
    return result


def _join_header_rows(
    header_rows: list[list[Any]], start_column: int, end_column: int
) -> list[str]:
    """Combine one or more header rows into stable unique column names."""
    width = end_column - start_column + 1
    filled_rows = [_filled_header_values(row, start_column, end_column) for row in header_rows]
    raw_columns: list[str] = []
    for offset in range(width):
        parts: list[str] = []
        for row in filled_rows:
            value = row[offset]
            if _is_empty(value):
                continue
            text = str(value).strip()
            if not parts or parts[-1] != text:
                parts.append(text)
        raw_columns.append(" — ".join(parts) if parts else f"Column {get_column_letter(start_column + offset)}")

    used: Counter[str] = Counter()
    result: list[str] = []
    for value in raw_columns:
        used[value] += 1
        result.append(value if used[value] == 1 else f"{value} ({used[value]})")
    return result


def _header_score(row: list[Any], next_row: list[Any] | None) -> int:
    """Heuristic score for candidate headers in messy business worksheets.

    The score intentionally rewards text labels and a populated data-like next row,
    while penalising report metadata (two-cell key/value rows) and obvious total
    rows. It is only used to locate table regions; the agent can always inspect a
    preview and select a table explicitly.
    """
    bounds = _row_bounds(row)
    if not bounds:
        return -100
    start, end = bounds
    values = [row[index - 1] for index in range(start, end + 1)]
    populated = [value for value in values if not _is_empty(value)]
    text_count = sum(isinstance(value, str) and bool(value.strip()) for value in populated)
    nontext_count = len(populated) - text_count
    score = len(populated) * 2 + text_count * 2 - nontext_count * 2
    if len(populated) <= 2:
        score -= 3
    first = str(populated[0]).strip().casefold() if populated else ""
    if first in {"total", "subtotal", "итого", "всего"}:
        score -= 8
    if next_row is not None:
        next_bounds = _row_bounds(next_row)
        if next_bounds:
            overlap = max(0, min(end, next_bounds[1]) - max(start, next_bounds[0]) + 1)
            score += min(overlap, 8)
            next_values = [next_row[index - 1] if index - 1 < len(next_row) else None for index in range(start, end + 1)]
            if any(not _is_empty(value) and not isinstance(value, str) for value in next_values):
                score += 2
    return score


def _looks_like_data_row(row: list[Any], header: list[Any]) -> bool:
    """Whether a row is plausibly data for the preceding header.

    Header rows are overwhelmingly labels, whereas business records often contain a
    number/date/bool or an identifier such as ``INV-100`` / ``CC-20``. This keeps
    an ordinary all-text record from becoming a second header only when it exhibits
    a data-like signal; otherwise a second header remains possible.
    """
    header_bounds = _row_bounds(header)
    row_bounds = _row_bounds(row)
    if not header_bounds or not row_bounds:
        return False
    start = max(header_bounds[0], row_bounds[0])
    end = min(header_bounds[1], row_bounds[1])
    if start > end:
        return False
    values = [row[index - 1] if index - 1 < len(row) else None for index in range(start, end + 1)]
    nonempty = [value for value in values if not _is_empty(value)]
    if not nonempty:
        return False
    if any(isinstance(value, (int, float, bool, datetime, date, time)) for value in nonempty):
        return True
    # Identifier-shaped strings are much more common in records than in headings.
    if any(
        isinstance(value, str)
        and bool(re.search(r"(?:\d|[-_/])", value.strip()))
        for value in nonempty
    ):
        return True
    return False


def _is_likely_metadata_pair(values: list[Any]) -> bool:
    """Recognise common report key/value labels before opening a narrow table."""
    first = next((value for value in values if not _is_empty(value)), None)
    if not isinstance(first, str):
        return False
    text = first.strip()
    folded = text.casefold()
    if folded in {
        "report", "entity", "run date", "currency", "confidential", "prepared by",
        "generated", "source", "period", "version", "owner",
    }:
        return True
    if text.startswith("*") or len(text) >= 80:
        return True
    nonempty = [value for value in values if not _is_empty(value)]
    if len(nonempty) == 2 and all(isinstance(value, str) and len(str(value).strip()) >= 60 for value in nonempty):
        return True
    return False


def _split_row_segments(row: list[Any]) -> list[tuple[int, int]]:
    """Find contiguous non-empty column islands in one worksheet row."""
    segments: list[tuple[int, int]] = []
    start: int | None = None
    for index, value in enumerate(row, start=1):
        if not _is_empty(value):
            if start is None:
                start = index
        elif start is not None:
            segments.append((start, index - 1))
            start = None
    if start is not None:
        segments.append((start, len(row)))
    return segments


def _is_group_header_row(row: list[Any], next_row: list[Any] | None) -> bool:
    """Identify a sparse first row of a merged/multi-row header."""
    if next_row is None:
        return False
    bounds = _row_bounds(row)
    next_bounds = _row_bounds(next_row)
    if not bounds or not next_bounds:
        return False
    own_count = _nonempty_count(row)
    next_count = _nonempty_count(next_row)
    if own_count < 1 or next_count <= own_count:
        return False
    # A merged heading has a gap *inside its own populated span*.  Do not include
    # trailing cells needed only by the next row: ``Prepared by | Operations``
    # followed by a five-column header is report metadata, not a grouped header.
    # The slice is end-exclusive, hence ``bounds[1]`` itself must be included.
    return any(_is_empty(value) for value in row[bounds[0] - 1 : bounds[1]])


def _is_total_row(row: list[Any], start_column: int, end_column: int) -> bool:
    """Return whether a row begins with a conventional total/subtotal label."""
    values = row[start_column - 1 : end_column]
    first = next((value for value in values if not _is_empty(value)), None)
    if not isinstance(first, str):
        return False
    return first.strip().casefold() in {"total", "subtotal", "grand total", "итого", "всего"}


def _column_index(columns: list[str], requested: str) -> int:
    if requested in columns:
        return columns.index(requested)
    folded = requested.strip().casefold()
    matches = [index for index, column in enumerate(columns) if column.casefold() == folded]
    if len(matches) == 1:
        return matches[0]
    raise ToolError("COLUMN_NOT_FOUND", f"Column {requested} not found", {"available_columns": columns})


def _table(state: dict[str, Any], table_id: str | None) -> dict[str, Any]:
    """Return a detected table while keeping model-facing ID recovery bounded.

    ``detect_tables`` creates opaque IDs. A tool-using model can safely omit or
    slightly corrupt that opaque value only when this session has precisely one
    detected table; there is then no data-selection ambiguity. Multi-table
    workbooks continue to require an exact server-issued ID.
    """
    tables = state.get("tables", {})
    raw = table_id.strip() if isinstance(table_id, str) else ""
    if raw:
        try:
            normalized = _opaque_id(raw, "tbl")
        except ToolError:
            normalized = ""
        if normalized and normalized in tables:
            return tables[normalized]
    if len(tables) == 1:
        return next(iter(tables.values()))
    if not raw:
        message = "table_id is required when more than one table is detected"
    else:
        message = f"Table {raw} not found"
    raise ToolError("TABLE_NOT_FOUND", message, {"available_table_ids": list(tables)})


def _rows_for_table(state: dict[str, Any], table: dict[str, Any]) -> Iterator[list[Any]]:
    for row_number, row in _iter_sheet_rows(state, table["sheet"], table["data_start_row"], table["end_row"]):
        del row_number
        values = row[table["start_column"] - 1 : table["end_column"]]
        if not all(_is_empty(value) for value in values):
            yield values


def _record(columns: list[str], row: list[Any]) -> dict[str, Any]:
    return {column: _json_value(row[index] if index < len(row) else None) for index, column in enumerate(columns)}


def _compact_rows(rows: list[dict[str, Any]], max_rows: int) -> list[dict[str, Any]]:
    return rows[: max(0, min(max_rows, MAX_PREVIEW_ROWS))]


@tool(_schema("workbook_introspect", "Returns compact workbook metadata and sheet dimensions. Call this before reading sheets.", {}))
def workbook_introspect(ctx: dict[str, Any], args: dict[str, Any]) -> dict[str, Any]:
    del args
    state = ctx["state"]
    sheets = _sheets(state)
    result = {"file_name": state["file_name"], "sheets": sheets}
    state["workbook_meta"] = result
    state["status"] = "introspected"
    return result


@tool(
    _schema(
        "sheet_preview",
        "Returns a small rectangular preview from one sheet; it never returns the whole sheet.",
        {
            "sheet": {"type": "string", "minLength": 1},
            "start_row": {"type": "integer", "minimum": 1, "default": 1},
            "max_rows": {"type": "integer", "minimum": 1, "maximum": 100, "default": 20},
            "max_columns": {"type": "integer", "minimum": 1, "maximum": 100, "default": 30},
        },
        ["sheet"],
    )
)
def sheet_preview(ctx: dict[str, Any], args: dict[str, Any]) -> dict[str, Any]:
    state = ctx["state"]
    sheet = args.get("sheet")
    if not isinstance(sheet, str) or not sheet:
        raise ToolError("INVALID_ARGUMENTS", "sheet must be a non-empty string")
    _ensure_sheet(state, sheet)
    start_row = args.get("start_row", 1)
    max_rows = args.get("max_rows", 20)
    max_columns = args.get("max_columns", 30)
    if not isinstance(start_row, int) or start_row < 1 or not isinstance(max_rows, int) or not 1 <= max_rows <= 100 or not isinstance(max_columns, int) or not 1 <= max_columns <= 100:
        raise ToolError("INVALID_ARGUMENTS", "Invalid preview bounds")

    rows: list[dict[str, Any]] = []
    for row_number, row in _iter_sheet_rows(state, sheet, start_row, start_row + max_rows - 1):
        values = [_json_value(value) for value in _trim_row(row, max_columns)]
        rows.append({"row_number": row_number, "values": values})
    return {"sheet": sheet, "start_row": start_row, "rows": rows}


def _candidate_has_data(candidate: dict[str, Any]) -> bool:
    return candidate["data_row_count"] > 0


def _is_numeric_cell(value: Any) -> bool:
    if isinstance(value, bool):
        return False
    return isinstance(value, (int, float, datetime, date, time))


def _append_candidate_data(candidate: dict[str, Any], row_number: int, row: list[Any]) -> None:
    """Record only bounded metadata for a detected data row.

    Detection must not retain every worksheet row: large workbooks are re-streamed by
    describe/query after the table coordinates are known. Keeping the previous record
    shape lets us safely bridge short visual blank gaps in O(table-width) memory.
    """
    start_column = candidate["start_column"]
    end_column = candidate["end_column"]
    values = row[start_column - 1 : end_column]
    if candidate["data_row_count"] == 0:
        candidate["data_start_row"] = row_number
    candidate["last_data_row"] = row_number
    candidate["last_data_values"] = values
    candidate["data_row_count"] += 1
    candidate["blank_row_count"] = 0
    numeric_cells = 0
    text_cells = 0
    long_text_cells = 0
    for value in values:
        if _is_empty(value):
            continue
        if _is_numeric_cell(value):
            numeric_cells += 1
            continue
        text_cells += 1
        if isinstance(value, str) and len(value.strip()) >= 80:
            long_text_cells += 1
    candidate["numeric_cells"] = candidate.get("numeric_cells", 0) + numeric_cells
    candidate["text_cells"] = candidate.get("text_cells", 0) + text_cells
    candidate["long_text_cells"] = candidate.get("long_text_cells", 0) + long_text_cells
    first = next((value for value in values if not _is_empty(value)), None)
    if isinstance(first, str) and first.lstrip().startswith("*"):
        candidate["star_rows"] = candidate.get("star_rows", 0) + 1
    if start_column > 1:
        occupancy = candidate.setdefault("left_occupancy", {})
        for column in range(1, start_column):
            stub = row[column - 1] if column - 1 < len(row) else None
            if not _is_empty(stub):
                occupancy[column] = occupancy.get(column, 0) + 1
        stub_index = start_column - 2
        stub = row[stub_index] if stub_index < len(row) else None
        if not _is_empty(stub):
            candidate["left_stub_populated"] = candidate.get("left_stub_populated", 0) + 1


def _looks_like_continuation_after_gap(
    row: list[Any], candidate: dict[str, Any], next_row: list[Any] | None
) -> bool:
    """Decide whether a non-empty row resumes a table after visual blank rows.

    The fast path recognises ordinary business records (numbers, dates, booleans or
    identifier-shaped values). For all-text records we additionally require a similar
    occupied-column shape to the preceding record. A likely fresh header followed by a
    record is deliberately rejected so neighbouring report blocks are not silently
    stitched together.
    """
    start_column = candidate["start_column"]
    end_column = candidate["end_column"]
    values = row[start_column - 1 : end_column]
    populated_count = _nonempty_count(values)
    width = end_column - start_column + 1
    minimum_cells = 1 if width == 1 else 2
    if populated_count < minimum_cells:
        return False

    header = candidate["header_rows"][-1][1][start_column - 1 : end_column]
    data_like = _looks_like_data_row(values, header)

    # ``Other | Value`` followed by ``x | 1`` after a gap is far more likely a
    # new narrow table than a continuation of the prior one. This still permits
    # records such as ``CC-30 | Support | 12000`` because they are data-like
    # against the original header.
    if not data_like and next_row is not None:
        next_values = next_row[start_column - 1 : end_column]
        label_count = sum(isinstance(value, str) and bool(value.strip()) for value in values if not _is_empty(value))
        if (
            label_count >= minimum_cells
            and _looks_like_data_row(next_values, values)
            and _header_score(values, next_values) >= 5
        ):
            return False

    if data_like:
        return True

    previous = candidate.get("last_data_values") or []
    previous_count = _nonempty_count(previous)
    # All-text data (names/categories) has no identifier or numeric signal. The
    # occupied shape is a lightweight conservative fallback; never bridge a
    # one-cell note into a multi-column table.
    return populated_count >= max(minimum_cells, math.ceil(previous_count * 0.6))


def _is_note_sheet_name(name: str) -> bool:
    folded = name.strip().casefold()
    return folded in _NOTE_SHEET_NAMES or folded.startswith("note")


def _is_preamble_sheet_name(name: str) -> bool:
    return name.strip().casefold() in _PREAMBLE_SHEET_NAMES


def _is_skipped_sheet_name(name: str) -> bool:
    return _is_note_sheet_name(name) or _is_preamble_sheet_name(name)


def _is_units_row(values: list[Any]) -> bool:
    nonempty = [value for value in values if not _is_empty(value)]
    if not nonempty:
        return False
    if any(_is_numeric_cell(value) for value in nonempty):
        return False
    unitish = 0
    for value in nonempty:
        if not isinstance(value, str):
            return False
        text = value.strip()
        if len(text) > 40:
            return False
        if _UNIT_HINT.search(text):
            unitish += 1
    return unitish >= max(1, math.ceil(len(nonempty) * 0.6))


def _candidate_is_notes(candidate: dict[str, Any]) -> bool:
    width = candidate["end_column"] - candidate["start_column"] + 1
    data_rows = candidate["data_row_count"]
    numeric = candidate.get("numeric_cells", 0)
    text = candidate.get("text_cells", 0)
    long_text = candidate.get("long_text_cells", 0)
    total = numeric + text
    sheet_name = str(candidate.get("sheet", ""))
    if _is_note_sheet_name(sheet_name) and (width <= 3 or (total and numeric / total < 0.2)):
        return True
    if _is_preamble_sheet_name(sheet_name) and total and numeric / total < 0.2:
        return True
    if width <= 2 and data_rows >= 1 and total:
        if long_text / max(text, 1) >= 0.35 and numeric / total < 0.25:
            return True
        if candidate.get("star_rows", 0) >= max(2, math.ceil(0.3 * data_rows)):
            return True
    return False


def _title_from_preamble(rows: Iterable[tuple[int, list[Any]]], limit: int = 400) -> str:
    parts: list[str] = []
    for _, row in rows:
        for value in row[:16]:
            if not isinstance(value, str):
                continue
            text = value.strip()
            if 2 <= len(text) <= 80:
                parts.append(text)
    seen: set[str] = set()
    unique: list[str] = []
    for part in parts:
        folded = part.casefold()
        if folded in seen:
            continue
        seen.add(folded)
        unique.append(part)
    return " ".join(unique)[:limit]


def _expand_stub_column(candidate: dict[str, Any]) -> None:
    """Include populated row-label/date columns immediately left of the header block."""
    start_column = candidate["start_column"]
    if start_column <= 1:
        return
    data_rows = candidate["data_row_count"]
    if data_rows < 1:
        return
    threshold = max(1, math.ceil(data_rows * 0.5))
    occupancy = dict(candidate.get("left_occupancy") or {})
    if not occupancy and candidate.get("left_stub_populated"):
        occupancy[start_column - 1] = int(candidate["left_stub_populated"])
    header_row = candidate["header_rows"][0][1]
    new_start = start_column
    scanned = 0
    for column in range(start_column - 1, 0, -1):
        scanned += 1
        if scanned > 8:
            break
        header = header_row[column - 1] if column - 1 < len(header_row) else None
        if not _is_empty(header):
            break
        if occupancy.get(column, 0) >= threshold:
            new_start = column
    candidate["start_column"] = new_start


def _absorb_units_header(state: dict[str, Any], candidate: dict[str, Any]) -> None:
    data_start = candidate.get("data_start_row")
    end_row = candidate.get("last_data_row")
    if not isinstance(data_start, int) or not isinstance(end_row, int) or data_start > end_row:
        return
    first_row: list[Any] | None = None
    for _, row in _iter_sheet_rows(state, candidate["sheet"], data_start, data_start):
        first_row = row
        break
    if first_row is None:
        return
    values = first_row[candidate["start_column"] - 1 : candidate["end_column"]]
    if not _is_units_row(values):
        return
    if data_start >= end_row:
        candidate["data_row_count"] = 0
        return
    candidate["data_start_row"] = data_start + 1
    candidate["data_row_count"] = max(0, candidate["data_row_count"] - 1)


def _compact_detected_table(table: dict[str, Any]) -> dict[str, Any]:
    return {
        "table_id": table["table_id"],
        "sheet": table["sheet"],
        "header_row": table["header_row"],
        "header_rows": table["header_rows"],
        "range": table["range"],
        "columns": list(table.get("columns") or []),
        "row_count": table.get("row_count"),
        "kind": table.get("kind", "data"),
        "title": table.get("title") or "",
    }


def _match_tokens(text: str) -> set[str]:
    words = re.findall(r"[^\W_]+", (text or "").casefold(), flags=re.UNICODE)
    return {word for word in words if len(word) >= 3 and word not in _MATCH_STOP_WORDS}


def _token_overlap(query_tokens: set[str], target_tokens: set[str]) -> set[str]:
    overlap: set[str] = set()
    for query_token in query_tokens:
        for target_token in target_tokens:
            if query_token == target_token or (
                len(query_token) >= 4
                and len(target_token) >= 4
                and (target_token.startswith(query_token) or query_token.startswith(target_token))
            ):
                overlap.add(query_token)
                break
    return overlap


def _sheet_alias_tokens(name: str) -> set[str]:
    tokens = _match_tokens(re.sub(r"[\s_\-]+", " ", name))
    parts = [part for part in re.split(r"[\s_\-]+", name.strip()) if part]
    if parts:
        suffix = parts[-1].casefold()
        if suffix in _SHEET_SUFFIX_ALIASES:
            tokens.update(_SHEET_SUFFIX_ALIASES[suffix])
    return tokens


def _negated_tokens(query: str) -> set[str]:
    tokens: set[str] = set()
    for match in re.finditer(
        r"(?:ignore|do not use|don't use|not(?: the)?)\s+([^.,;]+)",
        query,
        flags=re.IGNORECASE,
    ):
        tokens.update(_match_tokens(match.group(1)))
    return tokens


def _column_phrase_hits(query: str, columns: list[str]) -> int:
    folded = query.casefold()
    hits = 0
    for column in columns:
        label = column.strip()
        if len(label) < 4:
            continue
        if label.casefold() in folded:
            hits += 1
    return hits


def _score_text_overlap(query_tokens: set[str], target_tokens: set[str], negated: set[str]) -> int:
    overlap = _token_overlap(query_tokens, target_tokens)
    penalty = _token_overlap(negated, target_tokens)
    return 8 * len(overlap) - 20 * len(penalty)


def _ensure_sheet_labels(state: dict[str, Any]) -> dict[str, dict[str, Any]]:
    cached = state.get("sheet_labels")
    if isinstance(cached, dict) and cached:
        return cached
    labels: dict[str, dict[str, Any]] = {}
    for sheet in _sheets(state):
        name = sheet["name"]
        snippets: list[str] = []
        if sheet.get("state") != "hidden":
            for _, row in _iter_sheet_rows(state, name, 1, 8):
                for value in row[:12]:
                    if isinstance(value, str):
                        text = value.strip()
                        if 2 <= len(text) <= 80:
                            snippets.append(text)
        title = " ".join(dict.fromkeys(snippets))[:400]
        labels[name] = {
            "sheet": name,
            "title": title,
            "hidden": sheet.get("state") == "hidden",
            "tokens": sorted(_sheet_alias_tokens(name) | _match_tokens(title)),
        }
    state["sheet_labels"] = labels
    return labels


def _as_token_set(value: Any) -> set[str]:
    if isinstance(value, set):
        return {str(item) for item in value}
    if isinstance(value, list):
        return {str(item) for item in value if item}
    if isinstance(value, str) and value.strip():
        return _match_tokens(value)
    return set()


def _phrase_bonus(query: str, text: str) -> int:
    query_words = re.findall(r"[^\W_]+", (query or "").casefold(), flags=re.UNICODE)
    haystack = (text or "").casefold()
    bonus = 0
    for index in range(len(query_words) - 1):
        left, right = query_words[index], query_words[index + 1]
        if left in _MATCH_STOP_WORDS or right in _MATCH_STOP_WORDS:
            continue
        if f"{left} {right}" in haystack:
            bonus += 18
    return bonus


def _score_sheet(query: str, query_tokens: set[str], negated: set[str], label: dict[str, Any]) -> int:
    if label.get("hidden"):
        return -1000
    name = str(label["sheet"])
    title = str(label.get("title") or "")
    folded_query = query.casefold()
    score = _score_text_overlap(query_tokens, _as_token_set(label.get("tokens")), negated)
    score += _phrase_bonus(query, f"{name} {title}")
    if name.casefold() in folded_query:
        score += 50
    if _is_skipped_sheet_name(name) and not (_match_tokens(name) & query_tokens):
        score -= 40
    return score


def _score_table(query: str, query_tokens: set[str], negated: set[str], table: dict[str, Any], sheet_labels: dict[str, dict[str, Any]]) -> int:
    folded_query = query.casefold()
    sheet = str(table.get("sheet") or "")
    table_range = str(table.get("range") or "")
    columns = [str(column) for column in table.get("columns") or [] if column]
    label = sheet_labels.get(sheet) or {}
    tokens = _as_token_set(label.get("tokens")) | _sheet_alias_tokens(sheet) | _match_tokens(str(table.get("title") or label.get("title") or ""))
    score = _score_sheet(
        query,
        query_tokens,
        negated,
        {"sheet": sheet, "title": table.get("title") or label.get("title") or "", "tokens": sorted(tokens), "hidden": False},
    )
    locator = f"{sheet}!{table_range}".casefold()
    if locator in folded_query:
        score += 80
    if table_range.casefold() in folded_query:
        score += 40
    score += 12 * _column_phrase_hits(query, columns)
    score += _score_text_overlap(query_tokens, _match_tokens(" ".join(columns[:40])), negated) // 2
    if table.get("kind") == "notes":
        score -= 80
    return score


def _normalize_label(text: str) -> str:
    compact = re.sub(r"[\n\r]+", " ", text or "")
    compact = re.sub(r"[^\w\s]+", " ", compact, flags=re.UNICODE)
    return re.sub(r"\s+", " ", compact).strip().casefold()


def _is_date_stub_column(column: str) -> bool:
    folded = column.casefold()
    return folded.startswith("column ") or any(token in folded for token in ("date", "period", "month", "year", "time"))


def _phrase_in_text(phrase: str, text: str) -> bool:
    if not phrase or len(phrase) < 4:
        return False
    index = text.find(phrase)
    if index < 0:
        return False
    if index > 0 and text[index - 1].isalnum():
        return False
    end = index + len(phrase)
    if end < len(text) and text[end].isalnum():
        return False
    return True


def _column_series_label(column: str) -> str:
    parts = re.split(r"\s+[—–-]\s+", (column or "").strip())
    return parts[-1] if parts else (column or "")


def _column_phrase_in_query(query: str, column: str) -> bool:
    folded_query = _normalize_label(query)
    raw_query = query.casefold()
    for label_source in (column, _column_series_label(column)):
        label = _normalize_label(label_source)
        raw = re.sub(r"\s+", " ", (label_source or "").replace("\n", " ")).strip().casefold()
        if _phrase_in_text(label, folded_query) or _phrase_in_text(raw, raw_query):
            return True
    return False


def _column_tokens_covered(column: str, query_tokens: set[str]) -> bool:
    col_tokens = _match_tokens(_column_series_label(column)) or _match_tokens(column)
    if not col_tokens:
        return False
    distinctive = col_tokens - _GENERIC_COLUMN_TOKENS
    needed = distinctive or col_tokens
    return all(_token_overlap({token}, query_tokens) for token in needed)


def _suggested_select(query: str, columns: list[str]) -> list[str]:
    query_tokens = _match_tokens(query)
    matched: list[str] = []
    for column in columns:
        if _is_date_stub_column(column) and not _column_phrase_in_query(query, column):
            continue
        if _column_phrase_in_query(query, column) or (
            len(_match_tokens(column)) >= 2 and _column_tokens_covered(column, query_tokens)
        ):
            matched.append(column)
    if not matched:
        return []
    first = columns[0] if columns else ""
    if first and first not in matched and _is_date_stub_column(first):
        matched = [first, *matched]
    return matched


def _competing_columns(query: str, columns: list[str]) -> list[str]:
    query_tokens = _match_tokens(query)
    if not query_tokens:
        return []
    content_tokens = query_tokens - _GENERIC_COLUMN_TOKENS
    competitors: list[str] = []
    for column in columns:
        if _is_date_stub_column(column):
            continue
        col_tokens = _match_tokens(_column_series_label(column)) or _match_tokens(column)
        needle = content_tokens or query_tokens
        if _token_overlap(needle, col_tokens):
            competitors.append(column)
    return competitors


def _suggested_limit(query: str) -> int | None:
    match = _LATEST_COUNT_RE.search(query) or _CAP_COUNT_RE.search(query)
    if match:
        return max(1, min(int(match.group(1)), 10000))
    if _LATEST_PERIOD_RE.search(query):
        return 1
    return None


def _suggested_tail(query: str) -> bool:
    if _LATEST_COUNT_RE.search(query) or _LATEST_PERIOD_RE.search(query):
        return True
    return False


def _suggested_filters(query: str, columns: list[str]) -> list[dict[str, Any]]:
    if not _MISSING_QUERY_RE.search(query):
        return []
    targets = [column for column in columns if not _is_date_stub_column(column)]
    if not targets:
        return []
    folded = query.casefold()
    named = [column for column in targets if _column_phrase_in_query(query, column) or column.casefold() in folded]
    selected = named or (targets if len(targets) == 1 else [])
    return [{"field": column, "operator": "is_null"} for column in selected]


def _unique_winner(scored: list[tuple[int, Any]], min_score: int, gap: int) -> Any | None:
    if not scored:
        return None
    scored = sorted(scored, key=lambda item: item[0], reverse=True)
    best_score, best = scored[0]
    second = scored[1][0] if len(scored) > 1 else 0
    if best_score < min_score:
        return None
    if len(scored) == 1:
        return best
    needed = second + max(gap, int(second * 0.2))
    if best_score >= needed:
        return best
    return None


@tool(
    _schema(
        "detect_tables",
        "Detects tabular regions, including up to a bounded number of blank visual rows inside one table, and saves their IDs for later tools. Use optional sheet to restrict search. Returns compact table ids, ranges, column names and row counts — never the workbook.",
        {"sheet": {"type": "string", "minLength": 1}, "max_tables": {"type": "integer", "minimum": 1, "maximum": 80, "default": 10}},
    )
)
def detect_tables(ctx: dict[str, Any], args: dict[str, Any]) -> dict[str, Any]:
    state = ctx["state"]
    requested_sheet = args.get("sheet")
    max_tables = args.get("max_tables", 10)
    if requested_sheet is not None and (not isinstance(requested_sheet, str) or not requested_sheet):
        raise ToolError("INVALID_ARGUMENTS", "sheet must be a non-empty string")
    if not isinstance(max_tables, int) or not 1 <= max_tables <= 80:
        raise ToolError("INVALID_ARGUMENTS", "max_tables must be from 1 to 80")
    sheets = [requested_sheet] if requested_sheet else [item["name"] for item in _sheets(state)]
    if requested_sheet:
        _ensure_sheet(state, requested_sheet)

    try:
        max_internal_blank_rows = int(os.getenv("MAX_INTERNAL_BLANK_ROWS", "5"))
    except ValueError:
        max_internal_blank_rows = 5
    # A bounded threshold avoids scanning an entire whitespace-separated report
    # as one table while covering common paginated/visual spreadsheet gaps.
    max_internal_blank_rows = max(0, min(max_internal_blank_rows, 100))

    found: list[dict[str, Any]] = []

    def new_candidate(
        sheet: str, header_row_number: int, header_row: list[Any], start_column: int, end_column: int
    ) -> dict[str, Any]:
        return {
            "sheet": sheet,
            "header_rows": [(header_row_number, header_row)],
            "start_column": start_column,
            "end_column": end_column,
            "data_start_row": None,
            "last_data_row": None,
            "last_data_values": [],
            "data_row_count": 0,
            "blank_row_count": 0,
            "left_occupancy": {},
        }

    def open_candidates(
        sheet: str, row_number: int, row: list[Any], next_row: list[Any] | None
    ) -> list[dict[str, Any]]:
        """Open one or more table candidates from a potential header row."""
        bounds = _row_bounds(row)
        if bounds is None:
            return []
        if _is_group_header_row(row, next_row) and _nonempty_count(row) >= 2:
            assert next_row is not None
            next_bounds = _row_bounds(next_row)
            assert next_bounds is not None
            return [
                new_candidate(
                    sheet,
                    row_number,
                    row,
                    min(bounds[0], next_bounds[0]),
                    max(bounds[1], next_bounds[1]),
                )
            ]

        segments = _split_row_segments(row)
        candidate_segments = segments if len(segments) > 1 else [(bounds[0], bounds[1])]
        candidates: list[dict[str, Any]] = []
        for start_column, end_column in candidate_segments:
            segment = row[start_column - 1 : end_column]
            if len(segment) < 2 or _header_score(segment, None) < 4:
                continue
            if len(segment) == 2:
                if _is_likely_metadata_pair(segment):
                    continue
                following_segment = next_row[start_column - 1 : end_column] if next_row is not None else []
                if not _looks_like_data_row(following_segment, segment):
                    continue
            candidates.append(new_candidate(sheet, row_number, row, start_column, end_column))
        return candidates

    for sheet in sheets:
        # Stream each worksheet. Per active candidate we retain only two header rows,
        # one previous record shape and counters; workbook-height does not increase
        # detector memory use.
        iterator = iter(_iter_sheet_rows(state, sheet))
        current_item = next(iterator, None)
        active: list[dict[str, Any]] = []
        preamble: deque[tuple[int, list[Any]]] = deque(maxlen=8)
        while current_item is not None:
            row_number, row = current_item
            next_item = next(iterator, None)
            next_row = next_item[1] if next_item is not None else None
            bounds = _row_bounds(row)

            if bounds is None:
                next_active: list[dict[str, Any]] = []
                for candidate in active:
                    if not _candidate_has_data(candidate):
                        continue
                    candidate["blank_row_count"] += 1
                    if candidate["blank_row_count"] <= max_internal_blank_rows:
                        next_active.append(candidate)
                    else:
                        found.append(candidate)
                active = next_active
                if len(found) >= max_tables:
                    break
                current_item = next_item
                continue

            if not active:
                if bounds is not None:
                    preamble.append((row_number, row))
                active = open_candidates(sheet, row_number, row, next_row)
                title = _title_from_preamble(preamble)
                for candidate in active:
                    candidate["title"] = title
                current_item = next_item
                continue

            next_active = []
            restart_as_header = False
            for candidate in active:
                header_rows: list[tuple[int, list[Any]]] = candidate["header_rows"]
                overlap = max(
                    0,
                    min(candidate["end_column"], bounds[1])
                    - max(candidate["start_column"], bounds[0])
                    + 1,
                )
                if overlap == 0:
                    if _candidate_has_data(candidate):
                        found.append(candidate)
                        # The same physical row may be the header of a new block
                        # placed beside/after the prior table.  Reconsider it
                        # once the prior candidate has closed; otherwise a
                        # short, valid table such as ``Contact | Escalation``
                        # can be skipped merely because a blank gap kept the
                        # preceding table open.
                        restart_as_header = True
                    continue

                # A total/subtotal is a terminator even when it follows visual
                # spacing. Otherwise it could be mistaken for a text-shaped
                # continuation row and end up in the extracted data.
                if _candidate_has_data(candidate) and _is_total_row(
                    row, candidate["start_column"], candidate["end_column"]
                ):
                    found.append(candidate)
                    continue

                if _candidate_has_data(candidate) and candidate["blank_row_count"]:
                    if _looks_like_continuation_after_gap(row, candidate, next_row):
                        _append_candidate_data(candidate, row_number, row)
                        next_active.append(candidate)
                    else:
                        found.append(candidate)
                        restart_as_header = True
                    continue

                if not _candidate_has_data(candidate):
                    previous_header = header_rows[-1][1]
                    previous_has_merged_gap = any(
                        _is_empty(value)
                        for value in previous_header[candidate["start_column"] - 1 : candidate["end_column"]]
                    )
                    candidate_has_merged_gap = any(
                        _is_empty(value)
                        for value in row[candidate["start_column"] - 1 : candidate["end_column"]]
                    )
                    if (
                        len(header_rows) < 2
                        and _nonempty_count(row) >= 2
                        and _header_score(row, next_row) >= 5
                        and (
                            _is_group_header_row(previous_header, row)
                            or (
                                not _looks_like_data_row(row, previous_header)
                                and (previous_has_merged_gap or candidate_has_merged_gap)
                            )
                        )
                    ):
                        header_rows.append((row_number, row))
                        next_active.append(candidate)
                        continue

                _append_candidate_data(candidate, row_number, row)
                next_active.append(candidate)

            # If every active table was explicitly rejected as a continuation after
            # a short blank gap, revisit this row as a possible fresh header. That
            # prevents a neighbouring table from being swallowed or skipped.
            if restart_as_header and not next_active:
                active = open_candidates(sheet, row_number, row, next_row)
                title = _title_from_preamble(preamble)
                for candidate in active:
                    candidate["title"] = title
            else:
                active = next_active
            if len(found) >= max_tables:
                break
            current_item = next_item

        for candidate in active:
            if _candidate_has_data(candidate) and len(found) < max_tables:
                found.append(candidate)
        if len(found) >= max_tables:
            break

    existing_by_range: dict[tuple[str, str], dict[str, Any]] = {}
    for table in (state.get("tables") or {}).values():
        if isinstance(table, dict) and table.get("sheet") and table.get("range"):
            existing_by_range[(str(table["sheet"]), str(table["range"]))] = table
    response_tables: list[dict[str, Any]] = []
    for candidate in found:
        _expand_stub_column(candidate)
        _absorb_units_header(state, candidate)
        header_rows: list[tuple[int, list[Any]]] = candidate["header_rows"]
        if not _candidate_has_data(candidate) or _candidate_is_notes(candidate):
            continue
        columns = _join_header_rows(
            [row for _, row in header_rows], candidate["start_column"], candidate["end_column"]
        )
        data_start_row = candidate["data_start_row"]
        end_row = candidate["last_data_row"]
        assert isinstance(data_start_row, int) and isinstance(end_row, int)
        table_range = f"{get_column_letter(candidate['start_column'])}{header_rows[0][0]}:{get_column_letter(candidate['end_column'])}{end_row}"
        duplicate = existing_by_range.get((candidate["sheet"], table_range))
        if duplicate:
            response_tables.append(_compact_detected_table(duplicate))
            continue
        table_id = _id("tbl")
        persisted = {
            "table_id": table_id,
            "sheet": candidate["sheet"],
            "header_row": header_rows[0][0],
            "header_rows": [row_number for row_number, _ in header_rows],
            "data_start_row": data_start_row,
            "start_column": candidate["start_column"],
            "end_column": candidate["end_column"],
            "end_row": end_row,
            "range": table_range,
            "columns": columns,
            "row_count": candidate["data_row_count"],
            "kind": "data",
            "title": candidate.get("title") or "",
        }
        state.setdefault("tables", {})[table_id] = persisted
        existing_by_range[(persisted["sheet"], persisted["range"])] = persisted
        response_tables.append(_compact_detected_table(persisted))
    return {"tables": response_tables}


@tool(
    _schema(
        "match_tables",
        "Ranks detected tables against a natural-language request without reading the workbook into the model. Call this to pin one table or to see compact candidates. If no tables are saved yet, it inspects sheet titles first and only detects the matching sheet(s).",
        {
            "query": {"type": "string", "minLength": 1},
            "max_candidates": {"type": "integer", "minimum": 1, "maximum": 12, "default": 8},
            "sheet": {"type": "string", "minLength": 1},
        },
        ["query"],
    )
)
def match_tables(ctx: dict[str, Any], args: dict[str, Any]) -> dict[str, Any]:
    state = ctx["state"]
    query = args.get("query")
    max_candidates = args.get("max_candidates", 8)
    requested_sheet = args.get("sheet")
    if not isinstance(query, str) or not query.strip():
        raise ToolError("INVALID_ARGUMENTS", "query must be a non-empty string")
    if not isinstance(max_candidates, int) or not 1 <= max_candidates <= 12:
        raise ToolError("INVALID_ARGUMENTS", "max_candidates must be from 1 to 12")
    if requested_sheet is not None and (not isinstance(requested_sheet, str) or not requested_sheet):
        raise ToolError("INVALID_ARGUMENTS", "sheet must be a non-empty string")

    query_text = query.strip()
    query_tokens = _match_tokens(query_text)
    negated = _negated_tokens(query_text)
    sheet_labels = _ensure_sheet_labels(state)
    existing = [table for table in (state.get("tables") or {}).values() if isinstance(table, dict)]
    existing_sheets = {str(table.get("sheet")) for table in existing}

    target_sheets: list[str] = []
    if requested_sheet:
        _ensure_sheet(state, requested_sheet)
        target_sheets = [requested_sheet]
    else:
        scored_sheets = [
            (_score_sheet(query_text, query_tokens, negated, label), name)
            for name, label in sheet_labels.items()
            if not _is_skipped_sheet_name(name) or (_match_tokens(name) & query_tokens)
        ]
        winner = _unique_winner(scored_sheets, min_score=18, gap=10)
        ranked_sheets = sorted(scored_sheets, key=lambda item: item[0], reverse=True)
        positive = [name for score, name in ranked_sheets if score >= 12]
        visible_data = [
            name
            for name, label in sheet_labels.items()
            if not label.get("hidden") and not _is_skipped_sheet_name(name)
        ]
        if winner:
            target_sheets = [winner]
        elif 1 <= len(positive) <= 4:
            target_sheets = positive
        elif len(positive) > 4:
            sheet_candidates = [
                {
                    "sheet": name,
                    "title": sheet_labels.get(name, {}).get("title") or "",
                    "score": score,
                }
                for score, name in ranked_sheets[:max_candidates]
                if score > 0
            ]
            match = {
                "selected": None,
                "candidates": [],
                "sheet_candidates": sheet_candidates,
                "suggested_select": [],
                "ambiguous": True,
                "reason": "ambiguous_sheets",
            }
            state["table_match"] = {
                "query": query_text,
                "selected_table_id": None,
                "ambiguous": True,
                "reason": "ambiguous_sheets",
                "candidate_ids": [],
                "suggested_select": [],
            }
            return match
        elif existing:
            target_sheets = []
        elif len(visible_data) > 8:
            sheet_candidates = [
                {
                    "sheet": name,
                    "title": sheet_labels.get(name, {}).get("title") or "",
                    "score": score,
                }
                for score, name in ranked_sheets[:max_candidates]
            ]
            match = {
                "selected": None,
                "candidates": [],
                "sheet_candidates": sheet_candidates,
                "suggested_select": [],
                "ambiguous": True,
                "reason": "ambiguous_sheets",
            }
            state["table_match"] = {
                "query": query_text,
                "selected_table_id": None,
                "ambiguous": True,
                "reason": "ambiguous_sheets",
                "candidate_ids": [],
                "suggested_select": [],
            }
            return match
        else:
            target_sheets = visible_data

    for sheet in target_sheets:
        if sheet not in existing_sheets:
            detect_tables(ctx, {"sheet": sheet, "max_tables": 80})
            existing_sheets.add(sheet)

    if not (state.get("tables") or {}) and not target_sheets:
        detect_tables(ctx, {"max_tables": 80})

    tables = [
        table
        for table in (state.get("tables") or {}).values()
        if isinstance(table, dict) and table.get("kind") != "notes"
    ]
    scored_tables = [
        (_score_table(query_text, query_tokens, negated, table, sheet_labels), table)
        for table in tables
    ]
    scored_tables.sort(key=lambda item: item[0], reverse=True)
    selected_table = _unique_winner(scored_tables, min_score=16, gap=12)
    if selected_table is None and len(tables) == 1 and scored_tables and scored_tables[0][0] >= 0:
        selected_table = tables[0]
    candidates = [_compact_detected_table(table) for _, table in scored_tables[:max_candidates]]
    selected = _compact_detected_table(selected_table) if selected_table else None
    columns = list(selected_table.get("columns") or []) if selected_table else []
    suggested = _suggested_select(query_text, columns) if selected_table else []
    column_candidates: list[str] = []
    suggested_limit = _suggested_limit(query_text)
    suggested_tail = _suggested_tail(query_text)
    suggested_filters = _suggested_filters(query_text, suggested or columns)
    if selected is None and not tables:
        reason = "no_tables"
        ambiguous = False
    elif selected is None:
        reason = "ambiguous_tables"
        ambiguous = True
    else:
        phrase_hits = [column for column in columns if _column_phrase_in_query(query_text, column)]
        competitors = _competing_columns(query_text, columns)
        if not phrase_hits and len(competitors) >= 2:
            reason = "ambiguous_columns"
            ambiguous = True
            suggested = []
            suggested_filters = []
            column_candidates = competitors[:max_candidates]
        else:
            reason = "unique_table"
            ambiguous = False
    match = {
        "selected": selected,
        "candidates": candidates if selected is None else [selected],
        "sheet_candidates": [],
        "column_candidates": column_candidates,
        "suggested_select": suggested,
        "suggested_limit": suggested_limit,
        "suggested_tail": suggested_tail,
        "suggested_filters": suggested_filters,
        "ambiguous": ambiguous,
        "reason": reason,
    }
    state["table_match"] = {
        "query": query_text,
        "selected_table_id": selected["table_id"] if selected else None,
        "ambiguous": ambiguous,
        "reason": reason,
        "candidate_ids": [item["table_id"] for item in (candidates if selected is None else [selected])],
        "suggested_select": suggested,
        "suggested_limit": suggested_limit,
        "suggested_tail": suggested_tail,
        "suggested_filters": suggested_filters,
        "column_candidates": column_candidates,
    }
    return match

@tool(
    _schema(
        "describe_table",
        "Returns table columns, number of data rows and a small sample. Requires a table_id from detect_tables.",
        {"table_id": {"type": "string", "minLength": 1}, "sample_rows": {"type": "integer", "minimum": 0, "maximum": 20, "default": 5}},
        ["table_id"],
    )
)
def describe_table(ctx: dict[str, Any], args: dict[str, Any]) -> dict[str, Any]:
    table = _table(ctx["state"], args.get("table_id", ""))
    sample_rows = args.get("sample_rows", 5)
    if not isinstance(sample_rows, int) or not 0 <= sample_rows <= 20:
        raise ToolError("INVALID_ARGUMENTS", "sample_rows must be from 0 to 20")
    sample: list[dict[str, Any]] = []
    row_count = 0
    for row in _rows_for_table(ctx["state"], table):
        row_count += 1
        if len(sample) < sample_rows:
            sample.append(_record(table["columns"], row))
    return {
        "table_id": table["table_id"],
        "sheet": table["sheet"],
        "range": table["range"],
        "columns": table["columns"],
        "row_count": row_count,
        "sample_rows": sample,
    }


@tool(
    _schema(
        "list_column_values",
        "Returns bounded distinct values and counts for one column, useful to select exact filter values.",
        {"table_id": {"type": "string", "minLength": 1}, "column": {"type": "string", "minLength": 1}, "limit": {"type": "integer", "minimum": 1, "maximum": 500, "default": 50}},
        ["table_id", "column"],
    )
)
def list_column_values(ctx: dict[str, Any], args: dict[str, Any]) -> dict[str, Any]:
    state = ctx["state"]
    table = _table(state, args.get("table_id", ""))
    column = args.get("column")
    limit = args.get("limit", 50)
    if not isinstance(column, str) or not column or not isinstance(limit, int) or not 1 <= limit <= MAX_DISTINCT_VALUES:
        raise ToolError("INVALID_ARGUMENTS", "Invalid column or limit")
    index = _column_index(table["columns"], column)
    counts: Counter[str] = Counter()
    examples: dict[str, Any] = {}
    null_count = 0
    for row in _rows_for_table(state, table):
        value = row[index] if index < len(row) else None
        if _is_empty(value):
            null_count += 1
            continue
        safe_value = _json_value(value)
        key = json.dumps(safe_value, ensure_ascii=False, sort_keys=True)
        counts[key] += 1
        examples.setdefault(key, safe_value)
    ordered = sorted(counts, key=lambda key: (-counts[key], str(examples[key])))[:limit]
    return {
        "table_id": table["table_id"],
        "column": table["columns"][index],
        "values": [{"value": examples[key], "count": counts[key]} for key in ordered],
        "null_count": null_count,
        "distinct_count": len(counts),
        "truncated": len(counts) > limit,
    }


def _comparable(value: Any) -> Any:
    value = _json_value(value)
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        # LLM tool transports frequently serialize numeric filter values as
        # strings. Use one numeric representation so ``2000`` and "2000"
        # compare consistently without weakening boolean comparisons.
        return float(value)
    if isinstance(value, str):
        try:
            return float(value.replace(" ", "").replace(",", "."))
        except ValueError:
            return value.casefold()
    return str(value).casefold()


def _equal(left: Any, right: Any) -> bool:
    return _comparable(left) == _comparable(right)


def _matches_filter(value: Any, condition: dict[str, Any]) -> bool:
    operator = condition.get("operator", "eq")
    target = condition.get("value")
    if operator == "is_null":
        return _is_empty(value)
    if operator == "not_null":
        return not _is_empty(value)
    if operator in {"in", "not_in"}:
        if not isinstance(target, list):
            raise ToolError("INVALID_FILTER", f"Filter {operator} requires an array value")
        match = any(_equal(value, item) for item in target)
        return not match if operator == "not_in" else match
    if operator in {"contains", "not_contains"}:
        if target is None:
            raise ToolError("INVALID_FILTER", f"Filter {operator} requires value")
        match = str(_json_value(target)).casefold() in str(_json_value(value) or "").casefold()
        return not match if operator == "not_contains" else match
    if operator == "between":
        if not isinstance(target, list) or len(target) != 2:
            raise ToolError("INVALID_FILTER", "Filter between requires [lower, upper]")
        comparable = _comparable(value)
        return comparable is not None and _comparable(target[0]) <= comparable <= _comparable(target[1])
    if operator in {"eq", "neq"}:
        match = _equal(value, target)
        return not match if operator == "neq" else match
    if operator in {"gt", "gte", "lt", "lte"}:
        comparable = _comparable(value)
        target_value = _comparable(target)
        if comparable is None or target_value is None or type(comparable) is not type(target_value):
            return False
        return {"gt": comparable > target_value, "gte": comparable >= target_value, "lt": comparable < target_value, "lte": comparable <= target_value}[operator]
    raise ToolError("INVALID_FILTER", f"Unsupported filter operator {operator}", {"supported_operators": ["eq", "neq", "in", "not_in", "contains", "not_contains", "gt", "gte", "lt", "lte", "between", "is_null", "not_null"]})


def _prepare_query(table: dict[str, Any], args: dict[str, Any]) -> tuple[list[str], list[int], list[tuple[int, dict[str, Any]]]]:
    select = args.get("select") or table["columns"]
    if not isinstance(select, list) or not select or not all(isinstance(value, str) and value for value in select):
        raise ToolError("INVALID_ARGUMENTS", "select must be a non-empty list of column names")
    selected_indices = [_column_index(table["columns"], column) for column in select]
    selected_columns = [table["columns"][index] for index in selected_indices]
    filters = args.get("filters", [])
    if not isinstance(filters, list):
        raise ToolError("INVALID_FILTER", "filters must be an array")
    prepared_filters: list[tuple[int, dict[str, Any]]] = []
    for condition in filters:
        if not isinstance(condition, dict) or not isinstance(condition.get("field"), str):
            raise ToolError("INVALID_FILTER", "Each filter needs field and operator")
        prepared_filters.append((_column_index(table["columns"], condition["field"]), condition))
    return selected_columns, selected_indices, prepared_filters


def _iter_matching_records(state: dict[str, Any], table: dict[str, Any], selected_columns: list[str], selected_indices: list[int], filters: list[tuple[int, dict[str, Any]]]) -> Iterator[dict[str, Any]]:
    for row in _rows_for_table(state, table):
        if all(_matches_filter(row[index] if index < len(row) else None, condition) for index, condition in filters):
            yield {column: _json_value(row[index] if index < len(row) else None) for column, index in zip(selected_columns, selected_indices, strict=True)}


@tool(
    _schema(
        "query_table",
        "Filters a detected table and stores a bounded result set. For all matching rows use export_result afterwards.",
        {
            "table_id": {"type": "string", "minLength": 1},
            "select": {"type": "array", "items": {"type": "string"}, "minItems": 1},
            "filters": {"type": "array", "items": {"type": "object"}, "default": []},
            "limit": {"type": "integer", "minimum": 1, "maximum": 10000, "default": 200},
            "tail": {"type": "boolean", "default": False},
        },
        ["table_id"],
    )
)
def query_table(ctx: dict[str, Any], args: dict[str, Any]) -> dict[str, Any]:
    state = ctx["state"]
    table = _table(state, args.get("table_id", ""))
    limit = args.get("limit", 200)
    tail = args.get("tail", False)
    if not isinstance(limit, int) or not 1 <= limit <= 10000:
        raise ToolError("INVALID_ARGUMENTS", "limit must be from 1 to 10000")
    if not isinstance(tail, bool):
        raise ToolError("INVALID_ARGUMENTS", "tail must be a boolean")
    columns, indices, filters = _prepare_query(table, args)
    result_id = _id("res")
    relative_path = f"results/{result_id}.jsonl"
    output_path = session_file(state["session_id"], relative_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    row_count = 0
    stored_rows = 0
    preview_rows: list[dict[str, Any]] = []
    tail_buffer: deque[dict[str, Any]] | None = deque(maxlen=limit) if tail else None
    with output_path.open("w", encoding="utf-8") as output:
        for record in _iter_matching_records(state, table, columns, indices, filters):
            row_count += 1
            if tail_buffer is not None:
                tail_buffer.append(record)
                continue
            if stored_rows < limit:
                output.write(json.dumps(record, ensure_ascii=False, separators=(",", ":")) + "\n")
                stored_rows += 1
                if len(preview_rows) < MAX_QUERY_PREVIEW_ROWS:
                    preview_rows.append(record)
        if tail_buffer is not None:
            for record in tail_buffer:
                output.write(json.dumps(record, ensure_ascii=False, separators=(",", ":")) + "\n")
                stored_rows += 1
                if len(preview_rows) < MAX_QUERY_PREVIEW_ROWS:
                    preview_rows.append(record)
    state.setdefault("result_sets", {})[result_id] = {
        "result_id": result_id,
        "table_id": table["table_id"],
        "columns": columns,
        "row_count": row_count,
        "stored_rows": stored_rows,
        "preview_records": preview_rows,
        "truncated": row_count > stored_rows,
        "path": relative_path,
        "query": {"select": columns, "filters": args.get("filters", []), "limit": limit, "tail": tail},
    }
    return {
        "result_id": result_id,
        "columns": columns,
        "row_count": row_count,
        "stored_rows": stored_rows,
        "preview_rows": preview_rows,
        "truncated": row_count > stored_rows,
    }


@tool(
    _schema(
        "validate_result",
        "Validates a stored query result before finalization.",
        {"result_id": {"type": "string", "minLength": 1}, "required_columns": {"type": "array", "items": {"type": "string"}, "default": []}, "min_rows": {"type": "integer", "minimum": 0, "default": 1}},
        ["result_id"],
    )
)
def validate_result(ctx: dict[str, Any], args: dict[str, Any]) -> dict[str, Any]:
    state = ctx["state"]
    result_id = _opaque_id(args.get("result_id"), "res")
    result = state.get("result_sets", {}).get(result_id)
    if not result:
        raise ToolError("RESULT_NOT_FOUND", f"Result {result_id} not found")
    required_columns = args.get("required_columns", [])
    min_rows = args.get("min_rows", 1)
    if not isinstance(required_columns, list) or not all(isinstance(value, str) for value in required_columns) or not isinstance(min_rows, int) or min_rows < 0:
        raise ToolError("INVALID_ARGUMENTS", "Invalid required_columns or min_rows")
    missing = [column for column in required_columns if column not in result["columns"]]
    enough_rows = result["row_count"] >= min_rows
    valid = not missing and enough_rows
    # Validation is an explicit, server-side fact about this immutable query
    # result.  The n8n terminal safety-net and finalizer can therefore reject a
    # plausible-looking result that the agent never validated (or that failed
    # validation) instead of fabricating a successful extraction response.
    result["validation"] = {
        "valid": valid,
        "required_columns": list(required_columns),
        "min_rows": min_rows,
        "missing_columns": missing,
        "enough_rows": enough_rows,
        "row_count": result["row_count"],
    }
    return {"result_id": result["result_id"], "valid": valid, "row_count": result["row_count"], "columns": result["columns"], "missing_columns": missing, "min_rows": min_rows, "enough_rows": enough_rows}


def _result_or_raise(state: dict[str, Any], result_id: str) -> dict[str, Any]:
    result_id = _opaque_id(result_id, "res")
    result = state.get("result_sets", {}).get(result_id)
    if not result:
        raise ToolError("RESULT_NOT_FOUND", f"Result {result_id} not found")
    return result


def _csv_value(value: Any) -> Any:
    """Prevent spreadsheet formula execution when an exported CSV is opened."""
    safe = _json_value(value)
    if isinstance(safe, str) and safe.lstrip().startswith(("=", "+", "-", "@")):
        return f"'{safe}"
    return safe


def _export_csv(path: Path, records: Iterable[dict[str, Any]], columns: list[str]) -> int:
    count = 0
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=columns, extrasaction="ignore")
        writer.writerow({column: _csv_value(column) for column in columns})
        for record in records:
            writer.writerow({column: _csv_value(record.get(column)) for column in columns})
            count += 1
    return count


def _export_jsonl(path: Path, records: Iterable[dict[str, Any]]) -> int:
    count = 0
    with path.open("w", encoding="utf-8") as file:
        for record in records:
            file.write(json.dumps(record, ensure_ascii=False, separators=(",", ":")) + "\n")
            count += 1
    return count


@tool(
    _schema(
        "export_result",
        "Creates a downloadable CSV or JSONL artifact for every matching row of a query result.",
        {"result_id": {"type": "string", "minLength": 1}, "format": {"type": "string", "enum": ["csv", "jsonl"], "default": "csv"}},
        ["result_id"],
    )
)
def export_result(ctx: dict[str, Any], args: dict[str, Any]) -> dict[str, Any]:
    state = ctx["state"]
    result = _result_or_raise(state, args.get("result_id", ""))
    output_format = args.get("format", "csv")
    if output_format not in {"csv", "jsonl"}:
        raise ToolError("INVALID_ARGUMENTS", "format must be csv or jsonl")
    table = _table(state, result["table_id"])
    columns, indices, filters = _prepare_query(table, result["query"])
    artifact_id = _id("art")
    relative_path = f"artifacts/{artifact_id}.{output_format}"
    output_path = session_file(state["session_id"], relative_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    records = _iter_matching_records(state, table, columns, indices, filters)
    count = _export_csv(output_path, records, columns) if output_format == "csv" else _export_jsonl(output_path, records)
    state.setdefault("artifacts", {})[artifact_id] = {"artifact_id": artifact_id, "path": relative_path, "format": output_format, "row_count": count, "file_name": f"extraction.{output_format}"}
    return {"artifact_id": artifact_id, "format": output_format, "row_count": count}


@tool(
    _schema(
        "submit_clarification",
        "Saves one or more questions for a user when the workbook/request is ambiguous.",
        {"questions": {"type": "array", "minItems": 1, "maxItems": 10, "items": {"type": "object"}}},
        ["questions"],
    )
)
def submit_clarification(ctx: dict[str, Any], args: dict[str, Any]) -> dict[str, Any]:
    questions = args.get("questions")
    if not isinstance(questions, list) or not 1 <= len(questions) <= 10:
        raise ToolError("INVALID_ARGUMENTS", "questions must contain 1 to 10 entries")
    sanitized: list[dict[str, Any]] = []
    seen: set[str] = set()
    for question in questions:
        if not isinstance(question, dict) or not isinstance(question.get("id"), str) or not question["id"].strip() or not isinstance(question.get("question"), str) or not question["question"].strip():
            raise ToolError("INVALID_ARGUMENTS", "Every question needs non-empty id and question")
        if question["id"] in seen:
            raise ToolError("INVALID_ARGUMENTS", "Question ids must be unique")
        seen.add(question["id"])
        sanitized.append({key: _json_value(value) if key != "options" else [_json_value(item) for item in value] for key, value in question.items() if key in {"id", "question", "type", "options"}})
    token = _id("clr")
    state = ctx["state"]
    state.setdefault("clarifications", {})[token] = {"token": token, "status": "clarification_needed", "questions": sanitized, "answers": []}
    state["status"] = "clarification_needed"
    return {"token": token, "status": "clarification_needed", "questions": sanitized}
