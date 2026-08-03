"""Build a deterministic corpus of intentionally messy, real-world-shaped Excel files.

Run from the repository root with the Excel Tools image, for example:
  docker run --rm --user root -v "$PWD/excel-agent-tools/tests/fixtures/complex:/fixtures" \
    omegalul-excel-tools python /fixtures/generate_corpus.py

The files are synthetic and contain no customer data. They model failure-prone layouts
seen in exported financial/ERP/operations reports. Expected properties are recorded in
manifest.json and exercised by test_complex_corpus.py.
"""
from __future__ import annotations

import json
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).parent
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=14)


def save(workbook: Workbook, filename: str) -> None:
    path = ROOT / filename
    workbook.save(path)


def style_header(sheet, row: int, start: int, end: int) -> None:
    for cell in sheet[row][start - 1 : end]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def fixture_01() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"
    ws.merge_cells("A1:E1")
    ws["A1"] = "Northwind orders export"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A2:E2")
    ws["A2"] = "Reporting period: January 2026"
    ws["A4"] = "Prepared by"
    ws["B4"] = "Operations"
    ws.append(["Order ID", "Customer", "Amount", "Status", "Order date"])
    ws.append(["SO-1001", "Acme GmbH", 1250.5, "Paid", date(2026, 1, 5)])
    ws.append(["SO-1002", "Müller & Söhne", 980, "Draft", date(2026, 1, 7)])
    ws.append(["SO-1003", "ООО Север", 2110, "Paid", date(2026, 1, 9)])
    style_header(ws, 5, 1, 5)
    save(wb, "01_merged_title_and_preamble.xlsx")


def fixture_02() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"
    ws["A1"] = "Two independent tables share this sheet"
    ws.append(["Sales rep", "Region", "Revenue", None, "SKU", "Warehouse", "On hand"])
    ws.append(["Ada", "West", 1500, None, "SKU-1", "Paris", 42])
    ws.append(["Ben", "East", 900, None, "SKU-2", "Berlin", 17])
    ws.append(["Cara", "West", 2200, None, "SKU-3", "Paris", 0])
    style_header(ws, 2, 1, 3)
    style_header(ws, 2, 5, 7)
    save(wb, "02_two_side_by_side_tables.xlsx")


def fixture_03() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Multi level"
    ws.merge_cells("A1:B1")
    ws["A1"] = "Customer"
    ws.merge_cells("C1:D1")
    ws["C1"] = "FY2025 metrics"
    ws.append(["ID", "Name", "Revenue", "Margin"])
    ws.append(["C-01", "Atlas", 32000, 7800])
    ws.append(["C-02", "Boreal", 18000, 3100])
    ws.append(["C-03", "Citrus", 12500, 2900])
    style_header(ws, 1, 1, 4)
    style_header(ws, 2, 1, 4)
    save(wb, "03_merged_two_row_header.xlsx")


def fixture_04() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Late header"
    ws.append(["Report", "Aged receivables"])
    ws.append(["Entity", "Example Holdings"])
    ws.append(["Run date", date(2026, 2, 1)])
    ws.append(["Currency", "EUR"])
    ws.append(["Confidential", "Internal distribution only"])
    ws.append([])
    ws.append(["Invoice", "Debtor", "Due date", "Outstanding"])
    ws.append(["INV-1", "Fabrikam", date(2026, 1, 12), 760])
    ws.append(["INV-2", "Contoso", date(2026, 1, 21), 1200])
    ws.append(["INV-3", "Fabrikam", date(2026, 1, 29), 80])
    style_header(ws, 7, 1, 4)
    save(wb, "04_header_after_long_notes.xlsx")


def fixture_05() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Quarterly plan"
    ws.merge_cells("A1:A2")
    ws["A1"] = "Product"
    ws.merge_cells("B1:C1")
    ws["B1"] = "Q1"
    ws.merge_cells("D1:E1")
    ws["D1"] = "Q2"
    ws.append([None, "Units", "Revenue", "Units", "Revenue"])
    ws.append(["Gadget", 10, 1500, 12, 1900])
    ws.append(["Widget", 20, 2200, 21, 2400])
    ws.append(["Service", 5, 3000, 6, 3500])
    style_header(ws, 1, 1, 5)
    style_header(ws, 2, 1, 5)
    save(wb, "05_repeated_grouped_headers.xlsx")


def fixture_06() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Budget"
    ws.append(["Cost centre", "Owner", "Budget", "Actual"])
    ws.append(["CC-10", "Engineering", 50000, 48000])
    ws.append(["CC-20", "Marketing", 20000, 21500])
    ws.append([])
    ws.append([])
    ws.append([])
    ws.append(["CC-30", "Support", 12000, 9000])
    ws.append([])
    ws.append([])
    ws.append(["CC-40", "HR", 8000, 7800])
    ws.append(["Total", None, 90000, 86300])
    style_header(ws, 1, 1, 4)
    save(wb, "06_blank_line_and_total_row.xlsx")


def fixture_07() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Заказы 📦"
    ws.merge_cells("A1:D1")
    ws["A1"] = "Отчёт по заказам"
    ws.append(["Номер", "Контрагент", "Статус", "Сумма"])
    ws.append(["Z-1", "ООО Ромашка", "Оплачен", 15000])
    ws.append(["Z-2", "ООО Василёк", "Черновик", 800])
    ws.append(["Z-3", "Société Éclair", "Отгружен", 1900])
    style_header(ws, 2, 1, 4)
    hidden = wb.create_sheet("Internal calculations")
    hidden.sheet_state = "hidden"
    hidden.append(["Do not expose", "Value"])
    hidden.append(["Margin", 0.42])
    save(wb, "07_unicode_and_hidden_sheet.xlsx")


def fixture_08() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sparse layout"
    ws.merge_cells("A1:D1")
    ws["A1"] = "Operational KPI export"
    ws["A3"] = "Site"
    ws["B3"] = "Month"
    ws["C3"] = "Tickets"
    ws["D3"] = "SLA %"
    ws["F3"] = "Notes (not part of KPI table)"
    ws.append(["Lyon", "2026-01", 47, 0.94, None, "Late feed from vendor"])
    ws.append(["Paris", "2026-01", 65, 0.98, None, ""])
    ws.append(["Berlin", "2026-01", 31, 0.91, None, "Check mapping"])
    ws.append([])
    ws["F8"] = "Contact"
    ws["G8"] = "Escalation"
    ws["F9"] = "Ops desk"
    ws["G9"] = "P1"
    style_header(ws, 3, 1, 4)
    style_header(ws, 8, 6, 7)
    save(wb, "08_sparse_columns_and_side_notes.xlsx")


def fixture_09() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Duplicate columns"
    ws.append(["Region", "Sales", "Sales", "%"])
    ws.append(["North", 1100, 1300, 0.18])
    ws.append(["South", 900, 1000, 0.11])
    ws.append(["West", 1400, 1400, 0.22])
    style_header(ws, 1, 1, 4)
    save(wb, "09_duplicate_headers.xlsx")


def fixture_10() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Mixed types"
    ws.append(["Extract generated", "2026-03-01", None])
    ws.append([])
    ws.append(["Employee ID", "Start date", "Base salary", "Active", "Department"])
    ws.append(["E-100", date(2024, 6, 1), 70000, True, "Engineering"])
    ws.append(["E-101", date(2025, 2, 14), 55000, False, "Marketing"])
    ws.append(["E-102", date(2023, 10, 5), 63000, True, "Support"])
    ws.append(["E-103", date(2021, 1, 3), 82000, True, "Engineering"])
    style_header(ws, 3, 1, 5)
    save(wb, "10_dates_numbers_booleans.xlsx")


for make in (fixture_01, fixture_02, fixture_03, fixture_04, fixture_05, fixture_06, fixture_07, fixture_08, fixture_09, fixture_10):
    make()

manifest = {
    "description": "10 deterministic non-sensitive Excel fixtures built from commonly observed business-report layouts. They are intentionally awkward, not toy single-table sheets.",
    "source": "synthetic, locally generated by generate_corpus.py; each pattern is named so a public/customer workbook can be added later without changing expected behavior.",
    "fixtures": [
        {"file": "01_merged_title_and_preamble.xlsx", "patterns": ["merged_cells", "title", "preamble", "header_not_first_row"], "expected": [{"sheet": "Orders", "header_rows": [5], "columns": ["Order ID", "Customer", "Amount", "Status", "Order date"], "rows": 3}]},
        {"file": "02_two_side_by_side_tables.xlsx", "patterns": ["two_tables_one_sheet", "horizontal_split"], "expected": [{"sheet": "Dashboard", "header_rows": [2], "columns": ["Sales rep", "Region", "Revenue"], "rows": 3}, {"sheet": "Dashboard", "header_rows": [2], "columns": ["SKU", "Warehouse", "On hand"], "rows": 3}]},
        {"file": "03_merged_two_row_header.xlsx", "patterns": ["merged_cells", "two_row_header", "grouped_columns"], "expected": [{"sheet": "Multi level", "header_rows": [1, 2], "columns": ["Customer — ID", "Customer — Name", "FY2025 metrics — Revenue", "FY2025 metrics — Margin"], "rows": 3}]},
        {"file": "04_header_after_long_notes.xlsx", "patterns": ["header_after_notes", "preamble_rows_with_multiple_values"], "expected": [{"sheet": "Late header", "header_rows": [7], "columns": ["Invoice", "Debtor", "Due date", "Outstanding"], "rows": 3}]},
        {"file": "05_repeated_grouped_headers.xlsx", "patterns": ["merged_cells", "two_row_header", "repeated_group_headers"], "expected": [{"sheet": "Quarterly plan", "header_rows": [1, 2], "columns": ["Product", "Q1 — Units", "Q1 — Revenue", "Q2 — Units", "Q2 — Revenue"], "rows": 3}]},
        {"file": "06_blank_line_and_total_row.xlsx", "patterns": ["multiple_blank_rows_inside_table", "total_row"], "expected": [{"sheet": "Budget", "header_rows": [1], "columns": ["Cost centre", "Owner", "Budget", "Actual"], "rows": 4}]},
        {"file": "07_unicode_and_hidden_sheet.xlsx", "patterns": ["unicode_sheet_name", "hidden_sheet", "merged_title"], "expected": [{"sheet": "Заказы 📦", "header_rows": [2], "columns": ["Номер", "Контрагент", "Статус", "Сумма"], "rows": 3}]},
        {"file": "08_sparse_columns_and_side_notes.xlsx", "patterns": ["sparse_columns", "side_notes", "two_tables_one_sheet"], "expected": [{"sheet": "Sparse layout", "header_rows": [3], "columns": ["Site", "Month", "Tickets", "SLA %"], "rows": 3}, {"sheet": "Sparse layout", "header_rows": [8], "columns": ["Contact", "Escalation"], "rows": 1}]},
        {"file": "09_duplicate_headers.xlsx", "patterns": ["duplicate_headers"], "expected": [{"sheet": "Duplicate columns", "header_rows": [1], "columns": ["Region", "Sales", "Sales (2)", "%"], "rows": 3}]},
        {"file": "10_dates_numbers_booleans.xlsx", "patterns": ["dates", "numbers", "booleans", "header_not_first_row"], "expected": [{"sheet": "Mixed types", "header_rows": [3], "columns": ["Employee ID", "Start date", "Base salary", "Active", "Department"], "rows": 4}]},
    ],
}
(ROOT / "manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
print(f"Generated {len(manifest['fixtures'])} fixtures in {ROOT}")
