"""Public-workbook-shaped layouts: stub dates, unit rows, notes sheets, many regions."""
from __future__ import annotations

import io
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook


@pytest.fixture()
def client(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> TestClient:
    monkeypatch.setenv("SESSION_DIR", str(tmp_path / "sessions"))
    monkeypatch.setenv("API_KEY", "test-key")
    import importlib
    import app.main as main

    importlib.reload(main)
    return TestClient(main.app)


def _upload(client: TestClient, workbook: Workbook, name: str = "book.xlsx") -> str:
    stream = io.BytesIO()
    workbook.save(stream)
    response = client.post(
        "/api/v1/sessions",
        headers={"X-API-Key": "test-key"},
        files={"file": (name, stream.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"payload": '{"request":{"purpose":"messy layout test"}}'},
    )
    assert response.status_code == 201, response.text
    return response.json()["session_id"]


def _tool(client: TestClient, session_id: str, name: str, args: dict) -> dict:
    response = client.post(
        f"/api/v1/sessions/{session_id}/tool",
        headers={"X-API-Key": "test-key"},
        json={"name": name, "args": args},
    )
    assert response.status_code == 200, response.text
    payload = response.json()
    assert payload["ok"], payload
    return payload["result"]


def _pink_sheet_like() -> Workbook:
    workbook = Workbook()
    prices = workbook.active
    prices.title = "Monthly Prices"
    prices["A1"] = "World Bank Pink Sheet"
    prices["A2"] = "Monthly Prices"
    prices["B5"] = "Crude oil, Brent"
    prices["C5"] = "Crude oil, WTI"
    prices["D5"] = "Natural gas, Europe"
    prices["E5"] = "Banana, Europe"
    prices["B6"] = "($/bbl)"
    prices["C6"] = "($/bbl)"
    prices["D6"] = "($/mmbtu)"
    prices["E6"] = "($/kg)"
    prices["A7"] = "1960M01"
    prices["B7"] = 1.63
    prices["C7"] = "…"
    prices["D7"] = 0.4
    prices["E7"] = 0.14
    prices["A8"] = "2024M01"
    prices["B8"] = 80.1
    prices["C8"] = 74.2
    prices["D8"] = 9.5
    prices["E8"] = 0.9

    indices = workbook.create_sheet("Monthly Indices")
    indices["A1"] = "Monthly Indices"
    indices["F5"] = "Energy"
    indices["G5"] = "Non-energy"
    indices["F6"] = "(2010=100)"
    indices["G6"] = "(2010=100)"
    indices["A7"] = "1960M01"
    indices["F7"] = 12
    indices["G7"] = 18
    indices["A8"] = "2024M01"
    indices["F8"] = 40
    indices["G8"] = 22

    notes = workbook.create_sheet("Description")
    notes["A1"] = "* Crude oil"
    notes["B1"] = "Average of Dubai, Brent and WTI. Updated 2022 with methodological notes about interpolation."
    notes["A2"] = "* Natural gas"
    notes["B2"] = "Europe border price. See the 2019 revision for missing early observations and unit changes."
    notes["A3"] = "* Coal"
    notes["B3"] = "Australian thermal coal. Long prose with years such as 2020 and 2021 should not become a table."
    workbook.create_sheet("AFOSHEET")
    return workbook


def _regional_workbook() -> Workbook:
    workbook = Workbook()
    first = True
    for code, region, people in (
        ("neast_p", "North East", "People"),
        ("neast_w", "North East", "Women"),
        ("london_p", "London", "People"),
        ("yorks_p", "Yorkshire", "People"),
        ("nwest_p", "North West", "People"),
        ("emids_p", "East Midlands", "People"),
        ("wmids_p", "West Midlands", "People"),
        ("east_p", "East", "People"),
        ("wales_p", "Wales", "People"),
        ("scot_p", "Scotland", "People"),
    ):
        sheet = workbook.active if first else workbook.create_sheet(code)
        if first:
            sheet.title = code
            first = False
        sheet["A1"] = region
        sheet["A2"] = people
        sheet["B7"] = "Date"
        sheet["C7"] = "Employment rate (%)"
        sheet["D7"] = "Unemployment rate (%)"
        sheet["B8"] = "Jan 2024"
        sheet["C8"] = 72.1 if people == "People" else 68.4
        sheet["D8"] = 4.2
        sheet["B9"] = "Feb 2024"
        sheet["C9"] = 72.4 if people == "People" else 68.8
        sheet["D9"] = 4.1
    notes = workbook.create_sheet("Notes")
    notes["A1"] = "These notes explain sampling"
    notes["B1"] = "A long methodological paragraph from 2023 that must not be treated as an extractable regional table."
    return workbook


def test_detect_keeps_date_stub_skips_units_and_drops_notes(client: TestClient) -> None:
    session_id = _upload(client, _pink_sheet_like())
    detected = _tool(client, session_id, "detect_tables", {"max_tables": 20})
    sheets = {table["sheet"] for table in detected["tables"]}
    assert "Description" not in sheets
    assert "AFOSHEET" not in sheets
    prices = next(table for table in detected["tables"] if table["sheet"] == "Monthly Prices")
    assert prices["range"].startswith("A5:")
    indices = next(table for table in detected["tables"] if table["sheet"] == "Monthly Indices")
    assert indices["range"].startswith("A")
    assert "Crude oil, Brent" in prices["columns"]
    assert "Crude oil, WTI" in prices["columns"]
    assert any(column.lower().startswith("column a") or "1960" not in column for column in prices["columns"])
    assert prices["columns"][0] != "Crude oil, Brent"
    described = _tool(client, session_id, "describe_table", {"table_id": prices["table_id"], "sample_rows": 2})
    assert described["row_count"] == 2
    first = described["sample_rows"][0]
    assert "($/bbl)" not in first.values()
    assert first[prices["columns"][0]] == "1960M01"
    queried = _tool(
        client,
        session_id,
        "query_table",
        {
            "table_id": prices["table_id"],
            "select": [prices["columns"][0], "Crude oil, Brent", "Crude oil, WTI"],
            "filters": [{"field": "Crude oil, WTI", "operator": "is_null"}],
        },
    )
    assert queried["row_count"] == 1
    assert queried["preview_rows"][0]["Crude oil, Brent"] == 1.63


def test_match_tables_pins_named_sheet_and_keeps_employment_ambiguous(client: TestClient) -> None:
    pink_id = _upload(client, _pink_sheet_like(), "pink.xlsx")
    matched = _tool(
        client,
        pink_id,
        "match_tables",
        {"query": "From Monthly Prices, extract Crude oil, Brent and Crude oil, WTI. Ignore Description."},
    )
    assert matched["ambiguous"] is False
    assert matched["selected"]["sheet"] == "Monthly Prices"
    assert "Crude oil, Brent" in (matched["suggested_select"] or matched["selected"]["columns"])

    oil = _tool(client, pink_id, "match_tables", {"query": "Give me the oil price."})
    assert oil["ambiguous"] is True
    assert oil["reason"] == "ambiguous_columns"
    assert oil["selected"]["sheet"] == "Monthly Prices"
    assert "Crude oil, Brent" in oil["column_candidates"]
    assert "Crude oil, WTI" in oil["column_candidates"]

    gas = _tool(
        client,
        pink_id,
        "match_tables",
        {"query": "Extract Natural gas, Europe in $/mmbtu from the monthly prices table, last 12 observations."},
    )
    assert gas["ambiguous"] is False
    assert gas["suggested_select"][-1] == "Natural gas, Europe" or "Natural gas, Europe" in gas["suggested_select"]
    assert "Banana, Europe" not in gas["suggested_select"]
    assert gas["suggested_limit"] == 12
    assert gas["suggested_tail"] is True

    regional_id = _upload(client, _regional_workbook(), "regions.xlsx")
    north = _tool(
        client,
        regional_id,
        "match_tables",
        {"query": "From the North East people table, extract the unemployment rate (%). Use the People sheet, not Men or Women."},
    )
    assert north["ambiguous"] is False
    assert north["selected"]["sheet"] == "neast_p"
    assert any("unemployment" in column.casefold() for column in north["suggested_select"])
    assert not any(
        "employment rate" in column.casefold() and "unemployment" not in column.casefold()
        for column in north["suggested_select"]
    )

    women = _tool(
        client,
        regional_id,
        "match_tables",
        {"query": "North East women: extract Employment rate (%) for the latest period only."},
    )
    assert women["selected"]["sheet"] == "neast_w"
    assert women["suggested_limit"] == 1
    assert women["suggested_tail"] is True
    assert any("employment rate" in column.casefold() for column in women["suggested_select"])
    assert not any("unemployment" in column.casefold() for column in women["suggested_select"])

    ambiguous = _tool(client, regional_id, "match_tables", {"query": "What is the employment rate?"})
    assert ambiguous["ambiguous"] is True
    assert ambiguous["selected"] is None
    assert ambiguous["reason"] == "ambiguous_sheets"
    assert len(ambiguous["sheet_candidates"]) >= 3


def test_na_is_a_value_n_a_is_empty() -> None:
    from app.excel_tools import _is_empty

    assert _is_empty("n/a") is True
    assert _is_empty("N/A") is True
    assert _is_empty("…") is True
    assert _is_empty("null") is True
    assert _is_empty("NA") is False
    assert _is_empty("na") is False
    assert _is_empty("North America") is False


def test_na_region_code_survives_query(client: TestClient) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Regions"
    sheet["A1"] = "Region"
    sheet["B1"] = "Rate"
    sheet["A2"] = "NA"
    sheet["B2"] = 1.2
    sheet["A3"] = "EU"
    sheet["B3"] = 0.8
    session_id = _upload(client, workbook, "na.xlsx")
    detected = _tool(client, session_id, "detect_tables", {"max_tables": 5})
    table = detected["tables"][0]
    queried = _tool(
        client,
        session_id,
        "query_table",
        {"table_id": table["table_id"], "filters": [{"field": "Region", "operator": "eq", "value": "NA"}]},
    )
    assert queried["row_count"] == 1
    assert queried["preview_rows"][0]["Region"] == "NA"


def test_index_lookup_sheet_is_detected_not_skipped(client: TestClient) -> None:
    workbook = Workbook()
    lookup = workbook.active
    lookup.title = "Index"
    lookup["A1"] = "Code"
    lookup["B1"] = "Weight"
    lookup["A2"] = "WTI"
    lookup["B2"] = 0.4
    lookup["A3"] = "Brent"
    lookup["B3"] = 0.6
    notes = workbook.create_sheet("Description")
    notes["A1"] = "* notes"
    notes["B1"] = "A long methodological paragraph from 2023 that must not become a table."
    session_id = _upload(client, workbook, "index.xlsx")
    detected = _tool(client, session_id, "detect_tables", {"max_tables": 10})
    sheets = {table["sheet"] for table in detected["tables"]}
    assert "Index" in sheets
    assert "Description" not in sheets
    matched = _tool(client, session_id, "match_tables", {"query": "Extract weights from the Index sheet"})
    assert matched["selected"]["sheet"] == "Index"


def test_detect_tables_does_not_duplicate_the_same_range(client: TestClient) -> None:
    session_id = _upload(client, _pink_sheet_like())
    first = _tool(client, session_id, "detect_tables", {"sheet": "Monthly Prices", "max_tables": 10})
    second = _tool(client, session_id, "detect_tables", {"sheet": "Monthly Prices", "max_tables": 11})
    first_ids = [table["table_id"] for table in first["tables"]]
    second_ids = [table["table_id"] for table in second["tables"]]
    assert first_ids == second_ids
    ranges = [table["range"] for table in second["tables"]]
    assert len(ranges) == len(set(ranges))


def test_query_table_tail_returns_latest_rows(client: TestClient) -> None:
    session_id = _upload(client, _pink_sheet_like())
    detected = _tool(client, session_id, "detect_tables", {"sheet": "Monthly Prices"})
    table = detected["tables"][0]
    queried = _tool(
        client,
        session_id,
        "query_table",
        {
            "table_id": table["table_id"],
            "select": [table["columns"][0], "Crude oil, Brent"],
            "limit": 1,
            "tail": True,
        },
    )
    assert queried["stored_rows"] == 1
    assert queried["preview_rows"][0][table["columns"][0]] == "2024M01"
    missing = _tool(
        client,
        session_id,
        "query_table",
        {
            "table_id": table["table_id"],
            "select": [table["columns"][0], "Crude oil, WTI"],
            "filters": [{"field": "Crude oil, WTI", "operator": "is_null"}],
        },
    )
    assert missing["row_count"] == 1
    assert missing["preview_rows"][0]["Crude oil, WTI"] is None

