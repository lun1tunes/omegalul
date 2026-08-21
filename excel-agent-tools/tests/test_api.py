from __future__ import annotations

import io
import os
import shutil
import subprocess
import sys
import threading
import time
import zipfile
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from filelock import FileLock
from openpyxl import Workbook

from app.sessions import load_state, locked_session, session_lock_path


@pytest.fixture()
def client(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> TestClient:
    monkeypatch.setenv("SESSION_DIR", str(tmp_path / "sessions"))
    monkeypatch.setenv("API_KEY", "test-key")
    # Reload because production settings are read at module import.
    import importlib
    import app.main as main

    importlib.reload(main)
    return TestClient(main.app)


def workbook_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Заказы"
    sheet.append(["Заказ №", "Контрагент", "Сумма итого", "Статус"])
    sheet.append(["Z-1045", "ООО Ромашка", 15800.5, "Оплачен"])
    sheet.append(["Z-1046", "ООО Василек", 800, "Черновик"])
    sheet.append(["Z-1047", "ООО Ромашка", 1200, "Отгружен"])
    sheet.append([])
    sheet.append(["Справка", "не таблица"])
    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def test_direct_uvicorn_import_loads_service_local_env_with_safe_precedence(tmp_path: Path) -> None:
    """A native CMD launch must not depend on the .bat file to populate env."""
    service = tmp_path / "excel-agent-tools"
    shutil.copytree(Path(__file__).resolve().parents[1] / "app", service / "app")
    (service / "excel-tools.env").write_text(
        f"API_KEY=from-excel-tools-env\nSESSION_DIR={tmp_path / 'sessions'}\n",
        encoding="utf-8",
    )
    # The compatibility file must not override the preferred service file.
    (service / ".env").write_text("API_KEY=from-old-dot-env\n", encoding="utf-8")
    command = [
        sys.executable,
        "-c",
        "import app.main as main; print(main.API_KEY)",
    ]
    clean_env = os.environ.copy()
    clean_env.pop("API_KEY", None)
    clean_env["PYTHONPATH"] = str(service)
    loaded = subprocess.run(
        command,
        cwd=service,
        env=clean_env,
        check=True,
        capture_output=True,
        text=True,
    )
    assert loaded.stdout.strip() == "from-excel-tools-env"

    # Docker, a Windows service manager, or an explicit CMD ``set`` remains
    # authoritative over both files.
    process_env = dict(clean_env, API_KEY="from-process-environment")
    overridden = subprocess.run(
        command,
        cwd=service,
        env=process_env,
        check=True,
        capture_output=True,
        text=True,
    )
    assert overridden.stdout.strip() == "from-process-environment"


def upload(client: TestClient) -> str:
    response = client.post(
        "/api/v1/sessions",
        headers={"X-API-Key": "test-key"},
        files={"file": ("orders.xlsx", workbook_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"payload": '{"request":{"fields":["order_id"]}}'},
    )
    assert response.status_code == 201, response.text
    return response.json()["session_id"]


def tool(client: TestClient, session_id: str, name: str, args: dict) -> dict:
    response = client.post(f"/api/v1/sessions/{session_id}/tool", headers={"X-API-Key": "test-key"}, json={"name": name, "args": args})
    assert response.status_code == 200, response.text
    payload = response.json()
    assert payload["ok"], payload
    return payload["result"]


def test_full_excel_tool_flow_and_artifact(client: TestClient) -> None:
    assert client.get("/health").json() == {"ok": True}
    assert client.get("/api/v1/tools").status_code == 401
    schemas = client.get("/api/v1/tools", headers={"X-API-Key": "test-key"}).json()["tools"]
    names = {item["function"]["name"] for item in schemas}
    assert {"workbook_introspect", "sheet_preview", "detect_tables", "match_tables", "describe_table", "list_column_values", "query_table", "validate_result", "export_result", "submit_clarification", "resolve_clarification", "finalize_extraction"} <= names

    session_id = upload(client)
    meta = tool(client, session_id, "workbook_introspect", {})
    assert meta["sheets"][0]["name"] == "Заказы"
    preview = tool(client, session_id, "sheet_preview", {"sheet": "Заказы", "max_rows": 2})
    assert preview["rows"][1]["values"][0] == "Z-1045"
    detected = tool(client, session_id, "detect_tables", {"sheet": "Заказы"})
    table_id = detected["tables"][0]["table_id"]
    description = tool(client, session_id, "describe_table", {"table_id": table_id})
    assert description["columns"] == ["Заказ №", "Контрагент", "Сумма итого", "Статус"]
    statuses = tool(client, session_id, "list_column_values", {"table_id": table_id, "column": "статус"})
    assert statuses["distinct_count"] == 3
    queried = tool(client, session_id, "query_table", {"table_id": table_id, "select": ["Заказ №", "Контрагент", "Сумма итого"], "filters": [{"field": "Статус", "operator": "in", "value": ["Оплачен", "Отгружен"]}]})
    assert queried["row_count"] == 2
    assert queried["preview_rows"][0]["Заказ №"] == "Z-1045"
    validation = tool(client, session_id, "validate_result", {"result_id": queried["result_id"], "required_columns": ["Заказ №"]})
    assert validation["valid"] is True
    artifact = tool(client, session_id, "export_result", {"result_id": queried["result_id"], "format": "csv"})
    download = client.get(f"/api/v1/artifacts/{session_id}/{artifact['artifact_id']}", headers={"X-API-Key": "test-key"})
    assert download.status_code == 200
    assert "Z-1045" in download.content.decode("utf-8-sig")
    final = tool(client, session_id, "finalize_extraction", {"status": "success", "data": {"result_id": queried["result_id"], "row_count": 2, "returned_count": 2, "columns": queried["columns"], "records": queried["preview_rows"]}})
    assert final["output"]["status"] == "success"
    state = client.get(f"/api/v1/sessions/{session_id}/state", headers={"X-API-Key": "test-key"}).json()
    assert "file_path" not in state
    assert "path" not in state["artifacts"][artifact["artifact_id"]]
    assert state["result_sets"][queried["result_id"]]["validation"] == {
        "valid": True,
        "required_columns": ["Заказ №"],
        "min_rows": 1,
        "missing_columns": [],
        "enough_rows": True,
        "row_count": 2,
    }
    case_insensitive = tool(client, session_id, "validate_result", {"result_id": queried["result_id"], "required_columns": ["заказ №"]})
    assert case_insensitive["valid"] is True
    assert case_insensitive["missing_columns"] == []


def test_successful_finalization_requires_successful_result_validation(client: TestClient) -> None:
    session_id = upload(client)
    missing_result = client.post(
        f"/api/v1/sessions/{session_id}/tool",
        headers={"X-API-Key": "test-key"},
        json={"name": "finalize_extraction", "args": {"status": "success", "data": {}}},
    ).json()
    assert missing_result["ok"] is False
    assert missing_result["error"]["code"] == "INVALID_FINAL_OUTPUT"

    table_id = tool(client, session_id, "detect_tables", {"sheet": "Заказы"})["tables"][0]["table_id"]
    queried = tool(client, session_id, "query_table", {"table_id": table_id, "select": ["Заказ №"]})

    unvalidated = client.post(
        f"/api/v1/sessions/{session_id}/tool",
        headers={"X-API-Key": "test-key"},
        json={"name": "finalize_extraction", "args": {"status": "success", "data": {"result_id": queried["result_id"]}}},
    ).json()
    assert unvalidated["ok"] is False
    assert unvalidated["error"]["code"] == "RESULT_NOT_VALIDATED"

    rejected_validation = tool(
        client,
        session_id,
        "validate_result",
        {"result_id": queried["result_id"], "required_columns": ["Missing"]},
    )
    assert rejected_validation["valid"] is False
    failed_finalization = client.post(
        f"/api/v1/sessions/{session_id}/tool",
        headers={"X-API-Key": "test-key"},
        json={"name": "finalize_extraction", "args": {"status": "success", "data": {"result_id": queried["result_id"]}}},
    ).json()
    assert failed_finalization["ok"] is False
    assert failed_finalization["error"]["code"] == "RESULT_NOT_VALIDATED"

    accepted_validation = tool(client, session_id, "validate_result", {"result_id": queried["result_id"], "required_columns": ["Заказ №"]})
    assert accepted_validation["valid"] is True
    final = tool(client, session_id, "finalize_extraction", {"status": "success", "data": {"result_id": queried["result_id"]}})
    assert final["output"]["status"] == "success"


def test_structured_errors_batch_and_clarification(client: TestClient) -> None:
    session_id = upload(client)
    batch = client.post(
        f"/api/v1/sessions/{session_id}/tools/batch",
        headers={"X-API-Key": "test-key"},
        json={"calls": [{"call_id": "unknown", "name": "does_not_exist", "args": {}}, {"call_id": "introspect", "name": "workbook_introspect", "args": {}}]},
    )
    assert batch.status_code == 200
    results = batch.json()["results"]
    assert results[0]["ok"] is False and results[0]["error"]["code"] == "UNKNOWN_TOOL"
    assert results[1]["ok"] is True
    clarification = tool(client, session_id, "submit_clarification", {"questions": [{"id": "amount", "question": "Какую сумму использовать?", "type": "choice", "options": ["Сумма", "Сумма итого"]}]})
    answer = {"token": clarification["token"], "answers": [{"question_id": "amount", "answer": "Сумма итого"}]}
    resolved = tool(client, session_id, "resolve_clarification", answer)
    assert resolved == {"token": clarification["token"], "status": "resolved", "idempotent": False}
    # A retry after a client timeout is safe and does not mutate the answer.
    repeated = tool(client, session_id, "resolve_clarification", answer)
    assert repeated == {"token": clarification["token"], "status": "resolved", "idempotent": True}
    conflict = client.post(
        f"/api/v1/sessions/{session_id}/tool",
        headers={"X-API-Key": "test-key"},
        json={"name": "resolve_clarification", "args": {"token": clarification["token"], "answers": [{"question_id": "amount", "answer": "Сумма"}]}},
    ).json()
    assert conflict["ok"] is False and conflict["error"]["code"] == "CLARIFICATION_ALREADY_RESOLVED"


def test_agent_tool_transport_accepts_n8n_envelopes_and_top_level_arguments(client: TestClient) -> None:
    """n8n envelope variants stay safe without model-controlled session IDs.

    HTTP Request Tool 1.1 sends structured arguments as direct body fields;
    compatible older exports may use ``input``/``args``. The endpoint name fixes the tool and
    the session is supplied by the workflow, not selected by a model call.
    """
    session_id = upload(client)
    for body in (
        {"session_id": session_id, "input": {}},
        {"session_id": session_id, "args": {}},
    ):
        response = client.post(
            "/api/v1/agent-tools/workbook_introspect",
            headers={"X-API-Key": "test-key"},
            json=body,
        )
        assert response.status_code == 200, response.text
        assert response.json()["ok"] is True
    ambiguous = client.post(
        "/api/v1/agent-tools/workbook_introspect",
        headers={"X-API-Key": "test-key"},
        json={"session_id": session_id, "input": {}, "args": {}},
    )
    assert ambiguous.status_code == 422

    table_id = tool(client, session_id, "detect_tables", {"sheet": "Заказы"})["tables"][0]["table_id"]
    top_level = client.post(
        "/api/v1/agent-tools/describe_table",
        headers={"X-API-Key": "test-key"},
        json={"session_id": session_id, "table_id": table_id, "sample_rows": 1},
    )
    assert top_level.status_code == 200, top_level.text
    assert top_level.json()["result"]["table_id"] == table_id

    mixed = client.post(
        "/api/v1/agent-tools/describe_table",
        headers={"X-API-Key": "test-key"},
        json={"session_id": session_id, "input": {"table_id": table_id}, "sample_rows": 1},
    )
    assert mixed.status_code == 422


def test_n8n_http_tool_1_1_object_json_fields_are_normalized(client: TestClient) -> None:
    """n8n's JSON tool-parameter schema is object-only; API restores arrays."""
    session_id = upload(client)
    table_id = tool(client, session_id, "detect_tables", {"sheet": "Заказы"})["tables"][0]["table_id"]
    plan = client.post(
        "/api/v1/agent-tools/save_agent_plan",
        headers={"X-API-Key": "test-key"},
        json={
            "session_id": session_id,
            "plan": "Extract paid orders",
            "selected_table_ids": {"0": table_id},
            "select": {"0": "Заказ №", "1": "Статус"},
            "filters": {"0": {"field": "Статус", "operator": "eq", "value": "Оплачен"}},
            "assumptions": {},
            "warnings": {},
        },
    )
    assert plan.status_code == 200, plan.text
    assert plan.json()["ok"] is True
    queried = client.post(
        "/api/v1/agent-tools/query_table",
        headers={"X-API-Key": "test-key"},
        json={
            "session_id": session_id,
            "table_id": table_id,
            "select": {"0": "Заказ №", "1": "Статус"},
            "filters": {"0": {"field": "Статус", "operator": "eq", "value": "Оплачен"}},
        },
    )
    assert queried.status_code == 200, queried.text
    assert queried.json()["result"]["preview_rows"] == [{"Заказ №": "Z-1045", "Статус": "Оплачен"}]


def test_n8n_http_tool_optional_fields_remain_omitted(client: TestClient) -> None:
    """A minimal structured call must use tool defaults, not injected nulls."""
    session_id = upload(client)
    table_id = tool(client, session_id, "detect_tables", {"sheet": "Заказы"})["tables"][0]["table_id"]

    plan = client.post(
        "/api/v1/agent-tools/save_agent_plan",
        headers={"X-API-Key": "test-key"},
        json={"session_id": session_id, "plan": "Use the detected orders table"},
    )
    assert plan.status_code == 200, plan.text
    assert plan.json()["ok"] is True

    queried = client.post(
        "/api/v1/agent-tools/query_table",
        headers={"X-API-Key": "test-key"},
        json={"session_id": session_id, "table_id": table_id},
    )
    assert queried.status_code == 200, queried.text
    assert queried.json()["ok"] is True
    assert queried.json()["result"]["row_count"] == 3


def test_rejects_pathlike_or_non_excel_upload(client: TestClient) -> None:
    response = client.post("/api/v1/sessions", headers={"X-API-Key": "test-key"}, files={"file": ("../../bad.txt", b"not excel", "text/plain")})
    assert response.status_code == 415


def test_rejects_malformed_or_expanding_xlsx(client: TestClient, monkeypatch: pytest.MonkeyPatch) -> None:
    malformed = client.post(
        "/api/v1/sessions",
        headers={"X-API-Key": "test-key"},
        files={"file": ("fake.xlsx", b"not a zip archive", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert malformed.status_code == 415

    monkeypatch.setattr("app.main.MAX_ZIP_UNCOMPRESSED_SIZE", 10)
    archive = io.BytesIO()
    with zipfile.ZipFile(archive, "w") as zipped:
        zipped.writestr("[Content_Types].xml", "x" * 20)
        zipped.writestr("xl/workbook.xml", "x")
    expanding = client.post(
        "/api/v1/sessions",
        headers={"X-API-Key": "test-key"},
        files={"file": ("expanded.xlsx", archive.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert expanding.status_code == 413


def test_upload_schedules_ttl_cleanup_in_background(client: TestClient, monkeypatch: pytest.MonkeyPatch) -> None:
    calls: list[bool] = []
    monkeypatch.setattr("app.main.cleanup_expired_sessions", lambda: calls.append(True))
    upload(client)
    # TestClient waits for response background tasks, proving cleanup is scheduled
    # through FastAPI instead of running inline before session creation.
    assert calls == [True]


def test_expensive_discovery_is_cached_but_cache_is_not_exposed(
    client: TestClient, monkeypatch: pytest.MonkeyPatch
) -> None:
    """Repeated discovery returns stable IDs without reopening the workbook."""
    session_id = upload(client)
    import app.excel_tools as excel_tools
    import app.tools as tools

    original = excel_tools.detect_tables
    calls = 0

    def counted_detect(ctx: dict, args: dict) -> dict:
        nonlocal calls
        calls += 1
        return original(ctx, args)

    monkeypatch.setitem(tools.TOOL_FUNCS, "detect_tables", counted_detect)
    first = tool(client, session_id, "detect_tables", {"sheet": "Заказы"})
    second = tool(client, session_id, "detect_tables", {"sheet": "Заказы"})

    assert second == first
    assert calls == 1
    persisted = load_state(session_id)
    assert len(persisted["tool_cache"]) == 1
    public_state = client.get(
        f"/api/v1/sessions/{session_id}/state",
        headers={"X-API-Key": "test-key"},
    ).json()
    assert "tool_cache" not in public_state


def test_tool_calls_for_one_session_are_serialized(client: TestClient, monkeypatch: pytest.MonkeyPatch) -> None:
    """Regression test for stale load/save races between concurrent tool calls."""
    session_id = upload(client)
    import app.excel_tools as excel_tools

    original = excel_tools.workbook_introspect
    entered = threading.Event()
    release = threading.Event()

    def slow_introspect(ctx: dict, args: dict) -> dict:
        entered.set()
        assert release.wait(timeout=3)
        return original(ctx, args)

    monkeypatch.setitem(excel_tools.__dict__, "workbook_introspect", slow_introspect)
    # The registry is the function actually dispatched by execute_tool.
    import app.tools as tools
    monkeypatch.setitem(tools.TOOL_FUNCS, "workbook_introspect", slow_introspect)

    responses: list[int] = []

    def invoke(name: str) -> None:
        response = client.post(
            f"/api/v1/sessions/{session_id}/tool",
            headers={"X-API-Key": "test-key"},
            json={"name": name, "args": {}},
        )
        responses.append(response.status_code)

    first = threading.Thread(target=invoke, args=("workbook_introspect",))
    second = threading.Thread(target=invoke, args=("get_session_state",))
    first.start()
    assert entered.wait(timeout=3)
    second.start()
    time.sleep(0.1)
    # The second tool cannot complete before the first releases the session lock.
    assert responses == []
    release.set()
    first.join(timeout=3)
    second.join(timeout=3)
    assert responses == [200, 200]
    assert [item["tool"] for item in load_state(session_id)["tool_history"][-2:]] == [
        "workbook_introspect",
        "get_session_state",
    ]


def test_portable_session_lock_serializes_external_file_lock(client: TestClient) -> None:
    """The native Windows/Unix lock shares one persistent per-session lock path."""
    session_id = upload(client)
    acquired = threading.Event()
    lock_path = str(session_lock_path(session_id))

    with FileLock(lock_path):
        def acquire() -> None:
            with locked_session(session_id):
                acquired.set()

        contender = threading.Thread(target=acquire)
        contender.start()
        time.sleep(0.1)
        assert acquired.is_set() is False
    contender.join(timeout=3)
    assert acquired.is_set() is True


def test_opaque_tool_ids_allow_only_surrounding_whitespace(client: TestClient) -> None:
    """LLM tool calls may add a newline while copying server-generated IDs."""
    session_id = upload(client)
    table_id = tool(client, session_id, "detect_tables", {"sheet": "Заказы"})["tables"][0]["table_id"]
    description = tool(client, session_id, "describe_table", {"table_id": f"  {table_id}\n"})
    assert description["table_id"] == table_id
    queried = tool(client, session_id, "query_table", {"table_id": f"{table_id} ", "limit": 10})
    validated = tool(client, session_id, "validate_result", {"result_id": f"\n{queried['result_id']}\t"})
    assert validated["result_id"] == queried["result_id"]


def test_detector_stitches_bounded_blank_gaps_without_merging_new_blocks(
    client: TestClient, monkeypatch: pytest.MonkeyPatch
) -> None:
    """Several visual blank rows are ignored, but a longer gap starts a new table.

    This regression covers both record shapes seen in customer workbooks: numeric
    business rows and all-text rows.  The detector is deliberately bounded so a
    whitespace-separated report block cannot become part of the prior table.
    """
    monkeypatch.setenv("MAX_INTERNAL_BLANK_ROWS", "3")
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Gaps"
    # Three text columns exercise the all-text continuation path without
    # relaxing the existing two-cell metadata-pair protection.
    worksheet.append(["Category", "Status", "Channel"])
    worksheet.append(["Red", "Open", "Retail"])
    worksheet.append([])
    worksheet.append([])
    worksheet.append([])
    worksheet.append(["Blue", "Closed", "Partner"])
    # A new gap later in the same physical table must be evaluated independently:
    # accepting Blue resets the blank-gap counter before Green is encountered.
    worksheet.append([])
    worksheet.append([])
    worksheet.append(["Green", "Open", "Direct"])
    # Four rows exceed the configured three-row visual-gap cap. The following
    # header is a distinct block, not an extension of Category/Status.
    for _ in range(4):
        worksheet.append([])
    worksheet.append(["Metric", "Value"])
    worksheet.append(["Tickets", 42])
    stream = io.BytesIO()
    workbook.save(stream)

    response = client.post(
        "/api/v1/sessions",
        headers={"X-API-Key": "test-key"},
        files={"file": ("blank-gaps.xlsx", stream.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert response.status_code == 201, response.text
    session_id = response.json()["session_id"]

    tables = tool(client, session_id, "detect_tables", {"sheet": "Gaps", "max_tables": 10})["tables"]
    descriptions = [tool(client, session_id, "describe_table", {"table_id": table["table_id"], "sample_rows": 10}) for table in tables]
    stitched = next(
        description
        for description in descriptions
        if description["columns"] == ["Category", "Status", "Channel"]
    )
    separate = next(description for description in descriptions if description["columns"] == ["Metric", "Value"])
    assert stitched["row_count"] == 3
    assert stitched["sample_rows"] == [
        {"Category": "Red", "Status": "Open", "Channel": "Retail"},
        {"Category": "Blue", "Status": "Closed", "Channel": "Partner"},
        {"Category": "Green", "Status": "Open", "Channel": "Direct"},
    ]
    assert separate["row_count"] == 1
    assert separate["sample_rows"] == [{"Metric": "Tickets", "Value": 42}]


def test_csv_export_neutralizes_spreadsheet_formulas(client: TestClient, tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Export"
    worksheet.append(["Label", "Value"])
    # ``@`` remains a string in an .xlsx (unlike a real ``=`` formula, which
    # openpyxl deliberately returns as an uncomputed value with data_only=True).
    worksheet.append(["item-1", "@SUM(A1:A2)"])
    stream = io.BytesIO()
    workbook.save(stream)
    response = client.post(
        "/api/v1/sessions",
        headers={"X-API-Key": "test-key"},
        files={"file": ("formulas.xlsx", stream.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert response.status_code == 201
    session_id = response.json()["session_id"]
    table_id = tool(client, session_id, "detect_tables", {"sheet": "Export"})["tables"][0]["table_id"]
    result = tool(client, session_id, "query_table", {"table_id": table_id})
    artifact = tool(client, session_id, "export_result", {"result_id": result["result_id"], "format": "csv"})
    exported = client.get(f"/api/v1/artifacts/{session_id}/{artifact['artifact_id']}", headers={"X-API-Key": "test-key"})
    assert exported.status_code == 200
    text = exported.content.decode("utf-8-sig")
    assert "'@SUM(A1:A2)" in text

    from app.excel_tools import _csv_value, _export_csv

    assert _csv_value("=2+2") == "'=2+2"
    assert _csv_value("+1+1") == "'+1+1"
    assert _csv_value("-1+1") == "'-1+1"
    header_export = tmp_path / "header.csv"
    _export_csv(header_export, [{"=Formula header": "safe"}], ["=Formula header"])
    assert header_export.read_text(encoding="utf-8-sig").startswith("'=Formula header")


def test_n8n_session_endpoints_open_extract_and_tool_alias(client: TestClient) -> None:
    xlsx = Path("/home/lun1z/omegalul/simulation-model-example/golden-cases/golden_case_1/MONITORING_well_commissioning_dates.xlsx")
    if not xlsx.is_file():
        return
    opened = client.post(
        "/agent-tools/open_session",
        headers={"X-API-Key": "test-key"},
        json={
            "case_id": "CASE-1",
            "task_id": "TASK-1",
            "objective": "даты ввода",
            "inputs": {"excel_path": str(xlsx)},
        },
    )
    assert opened.status_code == 200, opened.text
    body = opened.json()
    assert body["ok"] is True
    session_id = body["session_id"]
    alias = client.post(
        "/agent-tools/detect_tables",
        headers={"X-API-Key": "test-key"},
        json={"session_id": session_id},
    )
    assert alias.status_code == 200, alias.text
    assert alias.json()["ok"] is True
    extracted = client.post(
        "/agent-tools/extract_commissioning",
        headers={"X-API-Key": "test-key"},
        json={"session_id": session_id},
    )
    assert extracted.status_code == 200, extracted.text
    result = extracted.json()
    assert result["status"] == "completed"
    wells = {item["well"] for item in result["data"]["facts"]}
    assert {"1601", "1602"} <= wells
    fetched = client.get(f"/sessions/{session_id}/result", headers={"X-API-Key": "test-key"})
    assert fetched.status_code == 200
    assert fetched.json()["status"] == "completed"
    assert client.post("/agent-tools/open_session", json={"objective": "даты"}).status_code == 401
